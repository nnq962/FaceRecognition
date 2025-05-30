import unicodedata
from flask import Flask, request, jsonify, send_file, render_template, redirect, url_for, session
import mimetypes
from datetime import datetime, timedelta
import pandas as pd
import os
import shutil
from insightface_detector import InsightFaceDetector
from utils.insightface_utils import process_image
import numpy as np
from flask_cors import CORS
from config import config
from annoy import AnnoyIndex
import faiss
import pickle
import utils.onvif_camera_tools as onvif_camera_tools
from utils.logger_config import LOGGER
from qr_code.generate_aruco_tags import generate_aruco_marker
import io
import cv2
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import tempfile
from functools import wraps
import bcrypt
from bson.objectid import ObjectId
import pillow_heif
from PIL import Image


app = Flask(__name__)
CORS(app)
# Thêm secret key cho session
app.secret_key = "YNK4FsxP7QdZ8tHu3BvT5jLrW9eG2mCa"  # Khóa bí mật cố định mạnh
INTERNAL_TOKEN = os.getenv("INTERNAL_TOKEN")
app.permanent_session_lifetime = timedelta(days=30)  # Thời gian mặc định của session permanent
detector = InsightFaceDetector()

# Kết nối tới MongoDB
user_collection = config.user_collection
camera_collection = config.camera_collection
admin_collection = config.admin_collection

# Define role hierarchy and permissions
ROLES_HIERARCHY = {
    'super_admin': ['admin', 'user'],  # Super admin can do everything
    'admin': ['user'],                 # Admin can do admin and user actions
    'user': []                         # User can only do user actions
}


# ----------------------------------------------------------------
# Define permissions for each API endpoint
API_PERMISSIONS = {
    # Admin management
    'get_admins': ['super_admin'],
    'add_admin': ['super_admin'],
    'get_admin': ['super_admin'],
    'update_admin': ['super_admin'],
    'delete_admin': ['super_admin'],
    
    # User management
    'get_users': ['user', 'admin', 'super_admin'],
    'add_user': ['admin', 'super_admin'],
    'get_user': ['admin', 'super_admin'],
    'update_user': ['admin', 'super_admin'],
    'delete_user': ['admin', 'super_admin'],
    'upload_photo': ['admin', 'super_admin'],
    'delete_photo': ['admin', 'super_admin'],
    'get_photos': ['admin', 'super_admin'],
    'view_photo': ['admin', 'super_admin'],
    'get_user_data': ['admin', 'super_admin'],
    
    # Attendance data
    'get_attendance': ['user', 'admin', 'super_admin'],
    'export_attendance': ['admin', 'super_admin'],
    'generate_excel': ['admin', 'super_admin'],
    'download': ['admin', 'super_admin'],
    'rebuild_all_users_embeddings': ['super_admin'],
    'get_qr_code': ['admin', 'super_admin'],

    # Camera
    'get_cameras': ['admin', 'super_admin'],
    'get_camera': ['admin', 'super_admin'],
    'add_camera': ['admin', 'super_admin'],
    'delete_camera': ['admin', 'super_admin'],
    'update_camera': ['admin', 'super_admin'],
    
    # Dashboard and basic operations
    'view' : ['user', 'admin', 'super_admin'],
    'export': ['admin', 'super_admin'],
    'users': ['admin', 'super_admin'],
    'manage_admins': ['super_admin'],
}

# ----------------------------------------------------------------
# Thêm hàm check_page_permission vào các template Jinja2
@app.context_processor
def utility_processor():
    return dict(check_page_permission=check_page_permission)


# ----------------------------------------------------------------
# Check if user has permission based on role
def has_permission(user_role, required_role):
    if user_role == required_role:
        return True
    
    # Check if user's role inherits the required role
    if required_role in ROLES_HIERARCHY.get(user_role, []):
        return True
    
    return False


# ----------------------------------------------------------------
# Role-based access decorator
def role_required(endpoint_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            token = request.headers.get("Internal-Token")
            if token == INTERNAL_TOKEN:
                return f(*args, **kwargs)

            # Nếu không phải request nội bộ thì kiểm tra session
            if 'user_id' not in session or 'role' not in session:
                return jsonify({
                    'status': 'error',
                    'message': 'Authentication required'
                }), 401
            
            user_role = session.get('role')
            required_roles = API_PERMISSIONS.get(endpoint_name, [])

            if not required_roles:
                return jsonify({
                    'status': 'error',
                    'message': 'No permissions defined for this endpoint'
                }), 403
            
            has_access = any(has_permission(user_role, role) for role in required_roles)
            if not has_access:
                return jsonify({
                    'status': 'error',
                    'message': 'Access denied: Insufficient permissions'
                }), 403

            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ----------------------------------------------------------------
# Helper function to apply RBAC to page routes as well
def check_page_permission(endpoint_name):
    # Check if user is logged in
    if 'user_id' not in session or 'role' not in session:
        return False
    
    # Get user's role from session
    user_role = session.get('role')
    
    # Get required roles for this endpoint
    required_roles = API_PERMISSIONS.get(endpoint_name, [])
    
    # Check if user has any of the required roles
    has_access = any(has_permission(user_role, role) for role in required_roles)
    
    return has_access


# ----------------------------------------------------------------
# Helper function to hash password
def hash_password(password):
    # Generate a salt and hash the password
    salt = bcrypt.gensalt()
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed_password.decode('utf-8')


# ----------------------------------------------------------------
# Helper function to check password
def check_password(password, hashed_password):
    return bcrypt.checkpw(password.encode('utf-8'), hashed_password.encode('utf-8'))


# ----------------------------------------------------------------
# Hàm decorator để bảo vệ các route yêu cầu đăng nhập
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Nếu gọi nội bộ với token hợp lệ → bypass login
        token = request.headers.get("Internal-Token")
        if token == INTERNAL_TOKEN:
            return f(*args, **kwargs)

        # Nếu không có user session → redirect về trang chủ
        if 'user_id' not in session:
            return redirect(url_for('home', next=request.url))

        return f(*args, **kwargs)
    return decorated_function


# ----------------------------------------------------------------
# Route đăng xuất
@app.route('/logout')
def logout():
    # Xóa session
    session.clear()
    return redirect(url_for('home'))


# ----------------------------------------------------------------
def generate_all_user_embeddings():
    """
    Cập nhật toàn bộ face_embeddings trong MongoDB,
    nếu thư mục hình ảnh của các users bị thay đổi.
    """
    users = user_collection.find()

    for user in users:
        user_id = user.get("user_id")
        photo_folder = user.get("photo_folder")

        LOGGER.info(f"Processing user_id: {user_id}...")

        if not photo_folder or not os.path.exists(photo_folder):
            LOGGER.warning(f"Photo folder does not exist for user_id {user_id}: {photo_folder}")
            continue

        face_embeddings = []
        photo_count = 0
        failed_photos = 0

        for file_name in os.listdir(photo_folder):
            file_path = os.path.join(photo_folder, file_name)

            if os.path.isfile(file_path):
                try:
                    # Gọi hàm xử lý để lấy face embedding và thông báo
                    face_embedding, processing_message = process_image(file_path, detector)

                    if face_embedding is not None:
                        face_embeddings.append({
                            "photo_name": file_name,
                            "embedding": face_embedding.tolist()
                        })

                        photo_count += 1
                        LOGGER.info(f"Added embedding for file {file_name} (user_id {user_id}): {processing_message}")
                    else:
                        failed_photos += 1
                        LOGGER.warning(f"Failed to process file {file_name} (user_id {user_id}): {processing_message}")

                except Exception as e:
                    failed_photos += 1
                    LOGGER.error(f"Error processing file {file_path}: {e}")

        # Cập nhật face_embeddings vào MongoDB theo user_id
        user_collection.update_one(
            {"user_id": user_id},
            {"$set": {"face_embeddings": face_embeddings}}
        )

        LOGGER.info(
            f"Completed processing user_id {user_id}. "
            f"Total photos processed: {photo_count}, Failed: {failed_photos}"
        )


# ----------------------------------------------------------------
def remove_accents(input_str):
    """
    Loại bỏ dấu tiếng Việt và chuyển Đ -> D, đ -> d
    """
    input_str = input_str.replace("Đ", "D").replace("đ", "d")  # Chuyển Đ -> D, đ -> d
    nfkd_form = unicodedata.normalize('NFKD', input_str)  # Chuẩn hóa Unicode
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])


# ----------------------------------------------------------------
def shorten_name(full_name):
    """
    Rút gọn tên thành dạng: N.N.Quyet (bỏ dấu và rút gọn)
    """
    full_name = remove_accents(full_name)  # Bỏ dấu trước khi rút gọn
    words = full_name.split()  # Tách tên theo khoảng trắng
    if len(words) >= 2:
        short_name = ".".join([w[0] for w in words[:-1]]) + "." + words[-1]
    else:
        short_name = words[0]  # Nếu chỉ có một từ, giữ nguyên
    return short_name


# ----------------------------------------------------------------
def build_faiss_index():
    """
    Tạo FAISS index (.faiss) và tệp ánh xạ index → user từ MongoDB.
    Lưu thông tin user_id, name và room_id trong tệp ánh xạ.
    """
    embeddings = []
    id_mapping = {}  # Lưu mapping từ FAISS index → user (user_id, name, room_id)
    index_counter = 0
    
    # Tạo thư mục chứa nếu chưa tồn tại
    os.makedirs(os.path.dirname(config.faiss_file), exist_ok=True)
    
    # Lấy thông tin từ MongoDB - tất cả người dùng
    users = user_collection.find(
        {}, 
        {"user_id": 1, "name": 1, "room_id": 1, "face_embeddings": 1}
    )
    
    user_count = 0
    for user in users:
        user_id = user.get("user_id")
        name = user.get("name", "Unknown")
        room_id = user.get("room_id")
        face_embeddings = user.get("face_embeddings", [])
        
        if not isinstance(face_embeddings, list) or not face_embeddings:
            LOGGER.warning(f"Bỏ qua user {user_id}: không có embeddings hoặc định dạng không đúng")
            continue
            
        user_count += 1
        for face_entry in face_embeddings:
            embedding = face_entry.get("embedding")
            if embedding and isinstance(embedding, list):
                try:
                    # Kiểm tra vector embedding hợp lệ
                    if len(embedding) < 128:  # Giả sử chiều vector là 128, điều chỉnh theo thực tế
                        LOGGER.warning(f"Bỏ qua embedding không hợp lệ của user {user_id}: độ dài {len(embedding)}")
                        continue
                        
                    embeddings.append(embedding)
                    id_mapping[index_counter] = {
                        "user_id": user_id,
                        "name": name,
                        "room_id": room_id
                    }
                    index_counter += 1
                except Exception as e:
                    LOGGER.error(f"Lỗi xử lý embedding cho user {user_id}: {str(e)}")
    
    if not embeddings:
        LOGGER.warning("Không có embeddings nào trong MongoDB!")
        return False
        
    # Convert sang numpy
    embeddings_array = np.array(embeddings, dtype=np.float32)
    vector_dim = embeddings_array.shape[1]
    
    try:
        # FAISS Inner Product index với chuẩn hóa vectors
        index = faiss.IndexFlatIP(vector_dim)
        index.add(embeddings_array)
        
        # Lưu FAISS index
        faiss.write_index(index, config.faiss_file)
        
        # Lưu ánh xạ index → user info
        with open(config.faiss_mapping_file, "wb") as f:
            pickle.dump(id_mapping, f)
            
        LOGGER.info(f"FAISS index đã được tạo và lưu vào {config.faiss_file}")
        LOGGER.info(f"Mapping index → user info đã được lưu vào {config.faiss_mapping_file}")
        LOGGER.info(f"Đã xử lý tổng cộng {index_counter} embeddings cho {user_count} người dùng")
        return True
    except Exception as e:
        LOGGER.error(f"Lỗi khi tạo hoặc lưu FAISS index: {str(e)}")
        return False


# ----------------------------------------------------------------
def build_ann_index():
    """
    Tạo tệp ann và tệp ánh xạ
    """
    embeddings = []
    id_mapping = {}  # Lưu mapping từ Annoy index → user (_id, full_name)

    index_counter = 0  # Đếm số embeddings

    # Lấy thông tin _id, full_name và embeddings từ MongoDB
    users = user_collection.find({}, {"_id": 1, "full_name": 1, "face_embeddings": 1})
    
    for user in users:
        user_id = user["_id"]
        full_name = user.get("full_name", "Unknown")  # Nếu không có full_name thì gán "Unknown"

        for face_entry in user.get("face_embeddings", []):  # Duyệt qua từng embedding
            embeddings.append(face_entry["embedding"])
            id_mapping[index_counter] = {
                "id": user_id,  # Sử dụng _id thay vì user_id
                "full_name": full_name  # Lưu full_name thay vì photo_name
            }
            index_counter += 1

    # Kiểm tra nếu không có embeddings
    if not embeddings:
        LOGGER.warning("Không có embeddings nào trong MongoDB!")
        return

    # Chuyển thành NumPy array để xử lý nhanh hơn
    embeddings = np.array(embeddings, dtype=np.float32)

    # Xác định số chiều của vector embedding
    vector_dim = len(embeddings[0])

    # Khởi tạo Annoy Index
    t = AnnoyIndex(vector_dim, 'euclidean')

    # Thêm từng embedding vào Annoy Index
    for i, emb in enumerate(embeddings):
        t.add_item(i, emb)

    # Xây dựng Annoy Index với 10 cây
    t.build(10)

    # Lưu file .ann
    t.save(config.ann_file)

    # Lưu id_mapping thành file .npy
    np.save(config.mapping_file, id_mapping)

    LOGGER.info(f"Annoy index has been successfully created and saved to {config.ann_file}.")
    LOGGER.info(f"Mapping index → user has been saved to {config.mapping_file}.")


# ----------------------------------------------------------------
@app.route('/api/get_users', methods=['GET'])
@login_required
@role_required('get_users')
def get_users():
    # Lấy tham số từ query
    room_id = request.args.get('room_id', None)
    without_face_embeddings = request.args.get('without_face_embeddings', '1') == '1'  # Mặc định là True

    # Tạo bộ lọc truy vấn MongoDB
    query = {}
    if room_id:
        query['room_id'] = room_id

    users = list(user_collection.find(query))

    for user in users:
        user.pop('_id', None)  # ❌ Bỏ _id không trả về
        if without_face_embeddings:
            user.pop('face_embeddings', None)

    return jsonify(users)


# ----------------------------------------------------------------
@app.route('/api/add_user', methods=['POST'])
@login_required
@role_required('add_user')
def add_user():
    data = request.json
    required_fields = ["name", "room_id", "user_id"]
    
    # Kiểm tra các trường bắt buộc
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing required fields"}), 400
    
    # Kiểm tra xem user_id đã tồn tại chưa
    existing_user = user_collection.find_one({"user_id": data["user_id"]})
    if existing_user:
        return jsonify({"error": "User code already exists"}), 400
    
    # Tạo đường dẫn thư mục cho user
    folder_path = f"{config.user_data_path}/{data['user_id']}"
    
    # Chuẩn bị dữ liệu để thêm vào MongoDB
    user = {
        "name": data["name"],
        "room_id": data["room_id"],
        "user_id": data["user_id"],
        "telegram_id": data.get("telegram_id"),
        "access_level": None,
        "created_at": config.get_vietnam_time(),
        "updated_at": None,
        "avatar_url": None,
        "face_embeddings": [],
        "photo_folder": folder_path,
        "active": True
    }

    try:
        # Thêm user vào MongoDB
        result = user_collection.insert_one(user)
        
        # Tạo thư mục cho user
        os.makedirs(folder_path, exist_ok=True)

        # Trả về thông tin user đã tạo
        return jsonify({
            "message": "User added successfully",
            "id": str(result.inserted_id),
            "user_id": data["user_id"],
            "photo_folder": folder_path
        }), 201
        
    except Exception as e:
        return jsonify({"error": f"Failed to add user: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/api/update_user/<user_id>', methods=['PUT'])
@login_required
@role_required('update_user')
def update_user(user_id):
    data = request.get_json()

    # Danh sách các trường được phép cập nhật
    allowed_fields = ['name', 'room_id', 'avatar_url', 'active']

    # Kiểm tra nếu có field không hợp lệ
    invalid_fields = [k for k in data.keys() if k not in allowed_fields]
    if invalid_fields:
        return jsonify({
            "error": "Some fields are not allowed to update",
            "invalid_fields": invalid_fields
        }), 400

    # Tạo dict chứa các trường cần cập nhật
    update_fields = {k: data[k] for k in allowed_fields if k in data}

    if not update_fields:
        return jsonify({"error": "No valid fields to update"}), 400

    # Cập nhật thời gian sửa đổi
    update_fields['updated_at'] = config.get_vietnam_time()

    # Cập nhật theo user_id
    result = user_collection.update_one(
        {"user_id": user_id},
        {"$set": update_fields}
    )

    if result.matched_count == 0:
        return jsonify({"error": "User not found"}), 404

    return jsonify({
        "message": "User updated successfully",
        "updated_fields": update_fields
    }), 200

# ----------------------------------------------------------------
@app.route('/api/delete_user/<user_id>', methods=['DELETE'])
@login_required
@role_required('delete_user')
def delete_user(user_id):
    try:
        # Tìm user theo user_id
        user = user_collection.find_one({"user_id": user_id})
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Lấy đường dẫn thư mục ảnh gốc
        folder_path = user.get("photo_folder")

        # Tạo đường dẫn backup
        def normalize_text(text):
            text = unicodedata.normalize("NFD", text)
            text = text.encode("ascii", "ignore").decode("utf-8")
            return text.replace(" ", "_").lower()

        name_slug = normalize_text(user.get("name", "unknown"))
        time_slug = config.get_vietnam_time().replace(":", "_").replace(" ", "_")
        backup_dir = os.path.expanduser("~/user_data_deleted")
        os.makedirs(backup_dir, exist_ok=True)

        # Tên thư mục backup: ~/user_data_deleted/nguyen_trung_lam_3hinc_2_2025_04_20_13_30_00
        backup_folder = os.path.join(
            backup_dir,
            f"{name_slug}_{user_id}_{time_slug}"
        )

        # Backup nếu tồn tại
        if folder_path and os.path.exists(folder_path):
            shutil.copytree(folder_path, backup_folder)

        # Xóa khỏi MongoDB
        result = user_collection.delete_one({"user_id": user_id})
        if result.deleted_count == 0:
            return jsonify({"error": "Failed to delete user"}), 500

        # Xóa thư mục gốc
        if folder_path and os.path.exists(folder_path):
            shutil.rmtree(folder_path)

        # Cập nhật lại chỉ mục FAISS
        build_faiss_index()

        return jsonify({
            "message": "User deleted and folder backed up successfully",
            "backup_folder": backup_folder
        }), 200

    except Exception as e:
        return jsonify({"error": f"Failed to delete user: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/api/get_user_data', methods=['GET'])
@login_required
@role_required('get_user_data')
def get_user_data():
    try:
        # Lấy tham số từ request
        user_id = request.args.get("user_id")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        camera_id = request.args.get("camera_id")
        room_id = request.args.get("room_id")  # Thêm tham số room_id
        
        # Kiểm tra tham số bắt buộc
        if not user_id or not start_date or not end_date or not camera_id:
            return jsonify({"error": "Missing required parameters"}), 400
            
        # Chọn collection dựa trên room_id
        if room_id:
            collection = config.database[f"{room_id}_logs"]
        else:
            collection = config.database['all_room_logs']
            
        # Chuyển đổi định dạng thời gian
        try:
            start_datetime = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
            end_datetime = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
            start_date_str = start_datetime.strftime("%Y-%m-%d")
            end_date_str = end_datetime.strftime("%Y-%m-%d")
            start_time_str = start_datetime.strftime("%H:%M:%S")
            end_time_str = end_datetime.strftime("%H:%M:%S")
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD HH:MM:SS"}), 400
            
        # Truy vấn dữ liệu từ MongoDB
        query = {
            "user_id": user_id,
            "date": {"$gte": start_date_str, "$lte": end_date_str},
            "timestamps": {
                "$elemMatch": {
                    "camera": camera_id,
                    "time": {"$gte": start_time_str, "$lte": end_time_str}
                }
            }
        }
        
        user_data = list(collection.find(
            query,
            {"_id": 0}  # Bỏ trường _id trong kết quả trả về
        ))
        
        if not user_data:
            return jsonify({"error": "No data found for given criteria"}), 404
            
        # Tùy chỉnh kết quả trả về (lọc timestamps nếu cần)
        filtered_data = []
        for record in user_data:
            # Chỉ giữ các timestamps khớp với camera_id và khoảng thời gian
            filtered_timestamps = [
                ts for ts in record["timestamps"]
                if ts["camera"] == camera_id
                and start_time_str <= ts["time"] <= end_time_str
            ]
            if filtered_timestamps:  # Chỉ thêm bản ghi nếu có timestamps khớp
                filtered_record = record.copy()
                filtered_record["timestamps"] = filtered_timestamps
                filtered_data.append(filtered_record)
                
        if not filtered_data:
            return jsonify({"error": "No matching timestamps found for given criteria"}), 404
            
        return jsonify(filtered_data), 200
        
    except Exception as e:
        return jsonify({"error": f"Failed to fetch user data: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/api/rebuild_all_users_embeddings', methods=['POST'])
@login_required
@role_required('rebuild_all_users_embeddings')
def rebuild_all_users_embeddings():
    try:
        # Chạy xử lý ảnh người dùng
        generate_all_user_embeddings()
        
        # Cập nhật ann index
        build_faiss_index()
        
        return jsonify({"message": "User photos processed and ann index updated"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to process and update: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/api/upload_photo/<user_id>', methods=['POST'])
@login_required
@role_required('upload_photo')
def upload_photo(user_id):
    if 'photo' not in request.files:
        return jsonify({"error": "No photo uploaded"}), 400

    photo = request.files['photo']
    original_filename = photo.filename

    # Kiểm tra xem file có phải là HEIC/HEIF không
    is_heic = False
    if original_filename.lower().endswith(('.heic', '.heif')):
        is_heic = True
        filename = os.path.splitext(original_filename)[0] + '.png'
    else:
        filename = original_filename

    # Lấy thông tin user theo user_id
    user = user_collection.find_one({"user_id": user_id})
    if not user:
        return jsonify({"error": "User not found"}), 404

    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    existing_photos = {entry["photo_name"] for entry in user.get("face_embeddings", []) or []}
    if filename in existing_photos:
        return jsonify({"error": "Photo already exists"}), 400

    try:
        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, filename)
        temp_path = os.path.join(folder_path, "temp_" + original_filename)

        photo.save(temp_path)

        # Chuyển HEIC → PNG nếu cần
        if is_heic:
            try:
                pillow_heif.register_heif_opener()
                img = Image.open(temp_path)
                img.save(file_path, format="PNG")
                os.remove(temp_path)
            except Exception as heic_error:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                return jsonify({"error": f"Failed to convert HEIC image: {str(heic_error)}"}), 400
        else:
            os.rename(temp_path, file_path)

        # Trích xuất face embedding
        face_embedding, processing_message = process_image(file_path, detector)

        if face_embedding is None:
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({"error": processing_message}), 400

        embedding_entry = {
            "photo_name": filename,
            "embedding": face_embedding.tolist()
        }

        user_collection.update_one(
            {"user_id": user_id},
            {"$push": {"face_embeddings": embedding_entry}}
        )

        build_faiss_index()

        return jsonify({
            "message": "Photo uploaded and face features saved",
            "photo_name": filename,
            "processing_details": processing_message + (" (Converted from HEIC to PNG)" if is_heic else ""),
            "converted": is_heic
        }), 200

    except Exception as e:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({"error": f"Failed to upload photo: {str(e)}"}), 500
    
    
# ----------------------------------------------------------------
@app.route('/api/delete_photo/<user_id>', methods=['DELETE'])
@login_required
@role_required('delete_photo')
def delete_photo(user_id):
    data = request.get_json()

    if not data or "file_name" not in data:
        return jsonify({"error": "File name is required"}), 400

    file_name = data["file_name"]

    # Tìm user theo user_id
    user = user_collection.find_one({"user_id": user_id})
    if not user:
        return jsonify({"error": "User not found"}), 404

    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    full_path = os.path.join(folder_path, file_name)

    if not os.path.exists(full_path):
        return jsonify({"error": "File not found"}), 404

    try:
        # Xóa ảnh thật khỏi thư mục
        os.remove(full_path)

        # Xóa embedding tương ứng trong MongoDB
        user_collection.update_one(
            {"user_id": user_id},
            {"$pull": {"face_embeddings": {"photo_name": file_name}}}
        )

        build_faiss_index()

        return jsonify({
            "message": "Photo and embedding deleted successfully"
        }), 200

    except Exception as e:
        return jsonify({
            "error": f"Failed to delete photo: {str(e)}"
        }), 500


# ----------------------------------------------------------------
@app.route('/api/get_photos/<user_id>', methods=['GET'])
@login_required
@role_required('get_photos')
def get_photos(user_id):
    # Tìm user theo user_id
    user = user_collection.find_one({"user_id": user_id})
    if not user:
        return jsonify({"error": "User not found"}), 404

    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    if not os.path.exists(folder_path):
        return jsonify({"error": "User photo folder does not exist"}), 404

    try:
        # Lấy danh sách ảnh trong thư mục
        photos = [
            f for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f))
        ]

        return jsonify({
            "photos": photos
        }), 200

    except Exception as e:
        return jsonify({
            "error": f"Failed to retrieve photos: {str(e)}"
        }), 500


# ----------------------------------------------------------------
@app.route('/api/view_photo/<user_id>/<filename>', methods=['GET'])
@login_required
@role_required('view_photo')
def view_photo(user_id, filename):
    # Lấy thông tin user từ MongoDB
    user = user_collection.find_one({"user_id": user_id})
    if not user:
        return jsonify({"error": "User not found"}), 404

    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    file_path = os.path.join(folder_path, filename)

    if not os.path.exists(file_path):
        return jsonify({"error": "Photo not found"}), 404

    try:
        # Tự động xác định loại ảnh
        mimetype = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
        return send_file(file_path, mimetype=mimetype)
    except Exception as e:
        return jsonify({"error": f"Failed to load photo: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/api/get_attendance', methods=['GET'])
@login_required
@role_required('get_attendance')
def get_attendance():
    try:
        # Lấy ngày hiện tại theo định dạng yyyy-mm-dd
        today = datetime.now().strftime("%Y-%m-%d")

        # Truy vấn tất cả bản ghi có trường date = hôm nay
        records = list(config.database['all_room_logs'].find({"date": today}))

        # Tạo danh sách kết quả
        attendance_list = []
        for record in records:
            attendance_list.append({
                "name": record.get("name", f"Unknown ({record.get('user_id')})"),
                "check_in_time": record.get("check_in_time", "N/A"),
                "check_out_time": record.get("check_out_time", "N/A"),
            })

        return jsonify(attendance_list), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/api/export_attendance', methods=['POST'])
@login_required
@role_required('export_attendance')
def export_attendance():
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        month = data.get("month")  # Định dạng YYYY-MM

        if not month:
            return jsonify({"error": "Thiếu thông tin month"}), 400
        
        # Lấy danh sách nhân viên từ collection users
        users = list(user_collection.find())
        results = []

        # Map ngày trong tuần sang tiếng Việt
        weekday_map = {
            0: "Thứ 2",
            1: "Thứ 3",
            2: "Thứ 4",
            3: "Thứ 5",
            4: "Thứ 6",
            5: "Thứ 7",
            6: "Chủ nhật"
        }

        # Duyệt qua từng nhân viên
        for user in users:
            user_id = user["user_id"]
            full_name = user["name"]

            # Lặp qua từng ngày trong tháng
            start_date = datetime.strptime(f"{month}-01", "%Y-%m-%d")
            end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

            current_date = start_date
            while current_date <= end_date:
                # Chỉ xét từ thứ 2 đến thứ 6
                if current_date.weekday() < 5:
                    date_str = current_date.strftime("%Y-%m-%d")
                    weekday_str = weekday_map[current_date.weekday()]

                    # Lấy dữ liệu từ mô hình mới
                    attendance_data = config.database['all_room_logs'].find_one({
                        "date": date_str,
                        "user_id": user_id
                    })
                    
                    # Tính giờ công và ghi chú
                    check_in, check_out, total_hours, note, kpi_deduction = calculate_work_hours_new(attendance_data)

                    # Thêm vào kết quả
                    results.append({
                        "ID": user_id,
                        "Tên nhân viên": full_name,
                        "Ngày": date_str,
                        "Thứ": weekday_str,
                        "Thời gian vào": check_in,
                        "Thời gian ra": check_out,
                        "Tổng giờ công": total_hours,
                        "Trừ KPI": kpi_deduction,
                        "Ghi chú": note
                    })

                # Sang ngày tiếp theo
                current_date += timedelta(days=1)

        # Trả về chỉ dữ liệu JSON, không tạo Excel
        return jsonify({
            "message": "Data retrieved successfully", 
            "data": results
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/api/generate_excel', methods=['POST'])
@login_required
@role_required('generate_excel')
def generate_excel():
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        edited_data = data.get("data", [])
        month_str = data.get("month", "")  # YYYY-MM
        
        if not edited_data:
            return jsonify({"error": "Không có dữ liệu để xuất"}), 400
        
        # Tạo đường dẫn tệp Excel động
        month_number = month_str.split("-")[1] if "-" in month_str else ""
        file_suffix = f"t{int(month_number)}" if month_number.isdigit() else "export"
        
        # Tạo tệp tạm thời
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, prefix=f'attendance_{file_suffix}_') as temp:
            temp_path = temp.name
        
        # Đảm bảo thứ tự cột đúng
        column_order = [
            "ID", 
            "Tên nhân viên", 
            "Ngày", 
            "Thứ", 
            "Thời gian vào", 
            "Thời gian ra", 
            "Tổng giờ công", 
            "Trừ KPI", 
            "Ghi chú"
        ]
        
        # Tạo DataFrame và sắp xếp cột theo thứ tự đã định
        df = pd.DataFrame(edited_data)
        
        # Đảm bảo tất cả các cột tồn tại (thêm cột thiếu nếu cần)
        for col in column_order:
            if col not in df.columns:
                df[col] = ""
        
        # Sắp xếp lại các cột theo thứ tự chỉ định
        df = df[column_order]
        
        # Xuất ra Excel với định dạng đẹp
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Chấm Công')
            
            # Lấy workbook và worksheet để định dạng
            workbook = writer.book
            worksheet = writer.sheets['Chấm Công']
            
            # Tạo định dạng cho tiêu đề
            header_font = Font(name='Calibri Light', bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Tạo định dạng cho nội dung
            content_font = Font(name='Calibri Light')
            content_alignment = Alignment(horizontal='center', vertical='center')
            
            # Áp dụng định dạng cho tiêu đề
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Áp dụng định dạng cho nội dung
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = content_font
                    cell.alignment = content_alignment
            
            # Tự động điều chỉnh chiều rộng cột
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Tạo đường viền cho bảng
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            # Áp dụng đường viền cho tất cả các ô có dữ liệu
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                        min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
        
        # Trả về đường dẫn đến tệp tạm thời để tải xuống
        return jsonify({
            "message": "Excel generated successfully", 
            "file": temp_path
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/api/download', methods=['GET'])
@login_required
@role_required('download')
def download_file():
    file_path = request.args.get('file')
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({"error": "File không tồn tại"}), 404
    
    # Lấy tên tệp từ đường dẫn
    filename = os.path.basename(file_path)
    
    # Gửi tệp về client
    response = send_file(file_path, as_attachment=True, download_name=filename)
    
    # Xóa tệp tạm thời sau khi gửi
    @response.call_on_close
    def cleanup():
        try:
            os.remove(file_path)
        except:
            pass
    
    return response


# ----------------------------------------------------------------
def calculate_work_hours_new(attendance_data):
    """Tính tổng giờ công trong ngày, ghi chú và trừ KPI theo cấu trúc dữ liệu mới."""
    if not attendance_data:  # Không có dữ liệu
        return None, None, None, "Nghỉ", ""
    
    check_in_time = attendance_data.get("check_in_time")
    check_out_time = attendance_data.get("check_out_time")
    
    if not check_in_time or not check_out_time:
        return check_in_time, check_out_time, None, "Nghỉ", ""
    
    # Chuyển đổi định dạng thời gian
    check_in_dt = datetime.strptime(check_in_time, "%H:%M:%S")
    check_out_dt = datetime.strptime(check_out_time, "%H:%M:%S")
    
    # Tính tổng giờ công (để đơn giản, giả sử cùng ngày)
    total_hours = round((check_out_dt - check_in_dt).total_seconds() / 3600, 2)
    
    # Thời gian chuẩn để so sánh
    start_time = datetime.strptime("08:00:00", "%H:%M:%S")
    late_threshold = datetime.strptime("08:10:00", "%H:%M:%S")
    noon_deadline = datetime.strptime("11:30:00", "%H:%M:%S")
    lunch_end = datetime.strptime("12:30:00", "%H:%M:%S")
    end_time = datetime.strptime("17:30:00", "%H:%M:%S")
    
    # Kiểm tra các điều kiện
    is_late = check_in_dt > late_threshold
    is_early_leave = check_out_dt < end_time
    missing_morning = check_in_dt > noon_deadline
    missing_afternoon = check_out_dt < lunch_end
    
    # Xác định ghi chú
    note = ""
    if missing_morning:
        note = "Thiếu công sáng"
    elif missing_afternoon:
        note = "Thiếu công chiều"
    elif is_late and is_early_leave:
        note = "Đi muộn & về sớm"
    elif is_late:
        note = "Đi muộn"
    elif is_early_leave:
        note = "Về sớm"
    
    # Xác định trừ KPI - chỉ trừ khi đi muộn quá 10 phút
    kpi_deduction = 1 if is_late else ""
    
    return check_in_time, check_out_time, total_hours, note, kpi_deduction
    

# ----------------------------------------------------------------
if config.init_database:
    LOGGER.warning("Initialize database")
    generate_all_user_embeddings()
    build_faiss_index()


# ----------------------------------------------------------------
@app.route('/api/get_cameras', methods=['GET'])
@login_required
@role_required('get_cameras')
def get_cameras():
    cameras = list(camera_collection.find({}, {"_id": 0}))
    return jsonify(cameras)


# ----------------------------------------------------------------
@app.route('/api/get_camera', methods=['GET'])
@login_required
@role_required('get_camera')
def get_camera():
    _id = request.args.get('_id')
    location = request.args.get('camera_location')

    if _id and location:
        return jsonify({"error": "Only filtering by _id or camera_location is allowed, not both."}), 400
    
    if _id:
        try:
            _id = int(_id)
        except ValueError:
            return jsonify({"error": "_id must be a number."}), 400
        camera = camera_collection.find_one({'_id': _id})
        if camera:
            return jsonify(camera)
        else:
            return jsonify({"error": "No camera found with the given _id."}), 404

    elif location:
        cameras = list(camera_collection.find({'camera_location': location}))
        if cameras:
            return jsonify(cameras)
        else:
            return jsonify({"error": "No cameras found with the given camera_location."}), 404
    
    else:
        return jsonify({"error": "Please provide either _id or camera_location for filtering."}), 400
    

# ----------------------------------------------------------------
@app.route('/api/add_camera', methods=['POST'])
@login_required
@role_required('add_camera')
def add_camera():
    data = request.get_json()

    # Lấy auto_discover từ query string (?auto_discover=0)
    auto_discover = request.args.get('auto_discover', '1') == '1'  # mặc định True

    # Các trường bắt buộc
    required_fields = ["camera_id", "name", "room_id"]
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing required fields: camera_id, name, room_id"}), 400

    camera_id = data["camera_id"]
    name = data["name"]
    room_id = data["room_id"]

    # Kiểm tra trùng camera_id
    if camera_collection.find_one({"camera_id": camera_id}):
        return jsonify({"error": "Camera code already exists."}), 400

    account = data.get("account")
    password = data.get("password")
    if not account or not password:
        return jsonify({"error": "Account and password are required."}), 400

    if auto_discover:
        # Không cho phép nhập IP và RTSP nếu auto_discover là True
        mac_address = data.get("MAC_address")
        if not mac_address:
            return jsonify({"error": "MAC_address is required when auto_discover is enabled."}), 400
        if "IP" in data or "RTSP" in data:
            return jsonify({"error": "Do not provide IP and RTSP when auto_discover is enabled."}), 400

        result = onvif_camera_tools.find_ip_and_rtsp_by_mac(mac_address, account, password)
        if not result:
            return jsonify({"error": f"Could not find device with MAC address {mac_address}."}), 404
        ip_address = result["IP"]
        rtsp_url = result["RTSP"]
    else:
        # Cho phép nhập đầy đủ IP/RTSP
        mac_address = data.get("MAC_address")
        ip_address = data.get("IP")
        rtsp_url = data.get("RTSP")

        if not all([mac_address, ip_address, rtsp_url]):
            return jsonify({"error": "MAC_address, IP, and RTSP are required when auto_discover is disabled."}), 400

    camera = {
        "camera_id": camera_id,
        "name": name,
        "room_id": room_id,
        "account": account,
        "password": password,
        "MAC_address": mac_address,
        "IP": ip_address,
        "RTSP": rtsp_url,
        "created_at": config.get_vietnam_time()
    }

    camera_collection.insert_one(camera)

    return jsonify({
        "message": "Camera added successfully."
    }), 201
    

# ----------------------------------------------------------------
@app.route('/api/delete_camera', methods=['DELETE'])
@login_required
@role_required('delete_camera')
def delete_camera():
    camera_id = request.args.get('camera_id')

    if not camera_id:
        return jsonify({"error": "camera_id is required for deletion."}), 400

    result = camera_collection.delete_one({'camera_id': camera_id})

    if result.deleted_count == 0:
        return jsonify({"error": "No camera found with the given camera_id."}), 404

    return jsonify({"message": "Camera deleted successfully."}), 200


# ----------------------------------------------------------------
@app.route('/api/update_camera', methods=['PUT'])
@login_required
@role_required('update_camera')
def update_camera():
    camera_id = request.args.get('camera_id')
    data = request.get_json()

    if not camera_id:
        return jsonify({"error": "camera_id is required for updating."}), 400

    if not data or not isinstance(data, dict):
        return jsonify({"error": "Invalid request body."}), 400

    # Loại bỏ trường không cho phép cập nhật
    update_fields = {key: data[key] for key in data if key != 'created_at'}

    if not update_fields:
        return jsonify({"error": "No valid fields provided for update."}), 400

    result = camera_collection.update_one(
        {'camera_id': camera_id},
        {'$set': update_fields}
    )

    if result.matched_count == 0:
        return jsonify({"error": "No camera found with the given camera_id."}), 404

    return jsonify({"message": "Camera updated successfully."}), 200


# ----------------------------------------------------------------
@app.route("/api/get_qr_code", methods=["GET"])
@login_required
@role_required('get_qr_code')
def get_qr_code():
    try:
        marker_id = int(request.args.get("id", 0))
        size = int(request.args.get("size", 400))
        dictionary_name = request.args.get("marker", "DICT_5X5_100")

        marker_image = generate_aruco_marker(dictionary_name, marker_id, size)

        # Encode ảnh thành PNG và đưa vào bộ nhớ RAM
        success, encoded_image = cv2.imencode(".png", marker_image)
        if not success:
            raise RuntimeError("Failed to encode marker image.")
        
        # Dùng BytesIO giả lập file
        image_io = io.BytesIO(encoded_image.tobytes())
        image_io.seek(0)

        return send_file(image_io, mimetype="image/png", as_attachment=False)

    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ----------------------------------------------------------------
# Get all admins
@app.route('/api/get_admins', methods=['GET'])
@role_required('get_admins')
@login_required
def get_admins():
    try:
        # Retrieve admins from collection, excluding password
        admins = list(admin_collection.find({}, {'password': 0}))
        
        # Convert ObjectId to string for JSON serialization
        for admin in admins:
            admin['_id'] = str(admin['_id'])
            admin['created_at'] = admin['created_at'].isoformat() if 'created_at' in admin else None
            admin['updated_at'] = admin['updated_at'].isoformat() if 'updated_at' in admin else None
            
        return jsonify({
            'status': 'success',
            'data': admins
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to retrieve admins: {str(e)}'
        }), 500


# ----------------------------------------------------------------
# Add a new admin
@app.route('/api/add_admin', methods=['POST'])
@login_required
@role_required('add_admin')
def add_admin():
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['username', 'password', 'full_name', 'role']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({
                    'status': 'error',
                    'message': f'Missing required field: {field}'
                }), 400
        
        # Check if username already exists
        existing_admin = admin_collection.find_one({'username': data['username']})
        if existing_admin:
            return jsonify({
                'status': 'error',
                'message': f'Username {data["username"]} already exists'
            }), 409
        
        # Hash password
        hashed_password = hash_password(data['password'])
        
        # Create admin document
        admin_data = {
            'username': data['username'],
            'password': hashed_password,
            'full_name': data['full_name'],
            'role': data['role'],
            'created_at': datetime.now(),
            'created_by': session.get('user_id')
        }
        
        # Insert into database
        result = admin_collection.insert_one(admin_data)
        
        return jsonify({
            'status': 'success',
            'message': 'Admin created successfully',
            'data': {
                'admin_id': str(result.inserted_id)
            }
        }), 201
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to create admin: {str(e)}'
        }), 500


# ----------------------------------------------------------------
# Get a specific admin
@app.route('/api/get_admin/<admin_id>', methods=['GET'])
@login_required
@role_required('get_admin')
def get_admin(admin_id):
    try:
        # Find admin by ID
        admin = admin_collection.find_one({'_id': ObjectId(admin_id)}, {'password': 0})
        
        if not admin:
            return jsonify({
                'status': 'error',
                'message': f'Admin not found with ID: {admin_id}'
            }), 404
        
        # Format data for response
        admin['_id'] = str(admin['_id'])
        admin['created_at'] = admin['created_at'].isoformat() if 'created_at' in admin else None
        admin['updated_at'] = admin['updated_at'].isoformat() if 'updated_at' in admin else None
        
        # Resolve updated_by field if it exists and contains an ObjectId
        if 'updated_by' in admin and admin['updated_by']:
            try:
                # Kiểm tra xem updated_by có phải là ObjectId không
                if ObjectId.is_valid(admin['updated_by']):
                    # Lấy thông tin người dùng đã cập nhật
                    updater = admin_collection.find_one({'_id': ObjectId(admin['updated_by'])}, {'_id': 1, 'username': 1, 'full_name': 1})
                    if updater:
                        admin['updated_by_name'] = updater.get('full_name') or updater.get('username')
            except:
                # Nếu có lỗi, giữ nguyên giá trị updated_by
                pass
        
        return jsonify({
            'status': 'success',
            'data': admin
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to retrieve admin: {str(e)}'
        }), 500


# ----------------------------------------------------------------
# Update an admin
@app.route('/api/update_admin/<admin_id>', methods=['PUT'])
@login_required
@role_required('update_admin')
def update_admin(admin_id):
    try:
        data = request.json
        
        # Find admin by ID
        existing_admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
        
        if not existing_admin:
            return jsonify({
                'status': 'error',
                'message': f'Admin not found with ID: {admin_id}'
            }), 404
        
        # Get current user information
        current_user_id = session.get('user_id')
        current_user = admin_collection.find_one({'_id': ObjectId(current_user_id)})
        
        # Prepare update data with current user information
        update_data = {
            'updated_at': datetime.now(),
            'updated_by': current_user_id,
            'updated_by_name': current_user.get('full_name', current_user.get('username', 'Unknown'))
        }
        
        # Update only provided fields
        fields_to_update = ['full_name', 'role']
        for field in fields_to_update:
            if field in data:
                update_data[field] = data[field]
        
        # Handle username change (check for duplicates)
        if 'username' in data and data['username'] != existing_admin['username']:
            check_username = admin_collection.find_one({'username': data['username']})
            if check_username:
                return jsonify({
                    'status': 'error',
                    'message': f'Username {data["username"]} already exists'
                }), 409
            update_data['username'] = data['username']
        
        # Handle password change (hash new password)
        if 'password' in data and data['password']:
            update_data['password'] = hash_password(data['password'])
        
        # Update in database
        if len(update_data) > 3:  # More than just updated_at, updated_by and updated_by_name
            admin_collection.update_one(
                {'_id': ObjectId(admin_id)},
                {'$set': update_data}
            )
            
            return jsonify({
                'status': 'success',
                'message': 'Admin updated successfully'
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': 'No valid fields to update'
            }), 400
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to update admin: {str(e)}'
        }), 500


# ----------------------------------------------------------------
# Delete an admin
@app.route('/api/delete_admin/<admin_id>', methods=['DELETE'])
@login_required
@role_required('delete_admin')
def delete_admin(admin_id):
    try:
        # Prevent deleting the currently logged-in admin
        current_user_id = session.get('user_id')
        if current_user_id == admin_id:
            return jsonify({
                'status': 'error',
                'message': 'Cannot delete the currently logged-in admin account'
            }), 403
        
        # Check if admin exists
        existing_admin = admin_collection.find_one({'_id': ObjectId(admin_id)})
        if not existing_admin:
            return jsonify({
                'status': 'error',
                'message': f'Admin not found with ID: {admin_id}'
            }), 404
        
        # Delete the admin
        result = admin_collection.delete_one({'_id': ObjectId(admin_id)})
        
        if result.deleted_count == 1:
            return jsonify({
                'status': 'success',
                'message': 'Admin deleted successfully'
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': 'Failed to delete admin'
            }), 500
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to delete admin: {str(e)}'
        }), 500


# ----------------------------------------------------------------
# User authentication API (login)
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        # Validate input
        if not username or not password:
            return jsonify({
                'status': 'error',
                'message': 'Username and password are required'
            }), 400
        
        # Find admin by username
        admin = admin_collection.find_one({'username': username})
        
        # Verify admin exists and password is correct
        if not admin or not check_password(password, admin['password']):
            return jsonify({
                'status': 'error',
                'message': 'Invalid username or password'
            }), 401
        
        # Create session
        session['user_id'] = str(admin['_id'])
        session['username'] = admin['username']
        session['full_name'] = admin.get('full_name', '')
        session['role'] = admin.get('role', '')
        
        # Return user info (exclude password)
        user_info = {
            'id': str(admin['_id']),
            'username': admin['username'],
            'full_name': admin.get('full_name', ''),
            'role': admin.get('role', '')
        }
        
        return jsonify({
            'status': 'success',
            'message': 'Authentication successful',
            'data': user_info
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Authentication failed: {str(e)}'
        }), 500


# ----------------------------------------------------------------
# User logout API
@app.route('/api/auth/logout', methods=['POST'])
@login_required
def api_logout():
    try:
        # Clear session
        session.clear()
        return jsonify({
            'status': 'success',
            'message': 'Logged out successfully'
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Logout failed: {str(e)}'
        }), 500


# ----------------------------------------------------------------
# Route đăng nhập
@app.route('/', methods=['GET', 'POST'])
def home():
    # Nếu đã đăng nhập, chuyển hướng đến trang chủ
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember') == 'on'  # Kiểm tra tùy chọn ghi nhớ đăng nhập
        
        # Kiểm tra thông tin đăng nhập
        admin = admin_collection.find_one({'username': username})
        
        # Sử dụng hàm check_password để kiểm tra mật khẩu đã mã hóa
        if admin and check_password(password, admin.get('password', '')):
            # Đăng nhập thành công
            session['user_id'] = str(admin['_id'])
            session['username'] = admin['username']
            session['full_name'] = admin.get('full_name', 'Admin')

            # Thêm dòng này để lưu vai trò vào session
            session['role'] = admin.get('role', 'user')  # Mặc định là 'user' nếu không có role
            
            # Nếu người dùng chọn "Ghi nhớ đăng nhập"
            if remember:
                # Mặc định, session sẽ hết hạn khi đóng trình duyệt
                # Thiết lập thời gian sống lâu hơn cho session (30 ngày)
                session.permanent = True
                app.permanent_session_lifetime = timedelta(days=30)
            else:
                # Nếu không ghi nhớ, session sẽ hết hạn khi đóng trình duyệt
                session.permanent = False
            
            # Mặc định chuyển hướng đến trang dashboard
            return redirect(url_for('dashboard'))
        else:
            # Đăng nhập thất bại
            error = "Tên đăng nhập hoặc mật khẩu không đúng!"
    
    # Hiển thị trang đăng nhập
    return render_template('login.html', error=error)


# ----------------------------------------------------------------
# Route cho trang theo dõi trạng thái
@app.route("/view")
@login_required
def view():
    # Kiểm tra quyền xem trang theo dõi trạng thái
    if not check_page_permission('view'):
        return render_template("error.html", message="Bạn không có quyền truy cập trang này"), 403
    
    return render_template("view.html")


# ----------------------------------------------------------------
# Route cho trang export dữ liệu
@app.route("/export")
@login_required
def export():
    # Kiểm tra quyền xem trang xuất dữ liệu
    if not check_page_permission('export'):
        return render_template("error.html", message="Bạn không có quyền truy cập trang này"), 403
    
    return render_template("export.html")


# ---------------------------------------------------------------- 
# Route cho quản lý người dùng
@app.route("/users")
@login_required
def users():
    # Kiểm tra quyền xem trang quản lý người dùng
    if not check_page_permission('users'):
        return render_template("error.html", message="Bạn không có quyền truy cập trang này"), 403
    
    # Truy vấn danh sách người dùng từ user_collection
    try:
        all_users = list(user_collection.find({}))
        # Chuyển ObjectId thành string
        for user in all_users:
            user['_id'] = str(user['_id'])
    except Exception as e:
        all_users = []
    
    return render_template("users.html", users=all_users)


# Trang lỗi khi không có quyền truy cập
@app.route("/error")
def error_page():
    message = request.args.get('message', 'Bạn không có quyền truy cập trang này')
    return render_template("error.html", message=message), 403


# ----------------------------------------------------------------
# Route cho trang dashboard (menu điều hướng)
@app.route("/dashboard")
@login_required
def dashboard():
    # Không cần kiểm tra quyền cho dashboard vì tất cả người dùng đã đăng nhập đều có thể xem
    return render_template("dashboard.html")


# ----------------------------------------------------------------
# Route cho trang quản lý admin
@app.route("/manage-admins")
@login_required
def manage_admins():
    # Kiểm tra quyền xem trang quản lý admin
    if not check_page_permission('manage_admins'):
        return render_template("error.html", message="Bạn không có quyền truy cập trang này"), 403
    
    # Lấy danh sách admin để hiển thị
    try:
        admins = list(admin_collection.find({}, {'password': 0}))
        # Chuyển ObjectId thành string
        for admin in admins:
            admin['_id'] = str(admin['_id'])
    except Exception as e:
        admins = []
    
    return render_template("manage_admins.html", admins=admins)


# ----------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=9621, debug=True)