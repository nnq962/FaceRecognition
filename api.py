import unicodedata
from flask import Flask, request, jsonify, send_file, render_template
import mimetypes
from datetime import datetime, timedelta
import pandas as pd
import os
import shutil
from insightface_detector import InsightFaceDetector
from insightface_utils import process_image
import numpy as np
from flask_cors import CORS
from gtts import gTTS
from config import config
from annoy import AnnoyIndex
import faiss
import pickle
import onvif_camera_tools
from utils.logger_config import LOGGER
from qr_code.generate_aruco_tags import generate_aruco_marker
import io
import cv2

app = Flask(__name__)
CORS(app)
detector = InsightFaceDetector()

# Kết nối tới MongoDB
users_collection = config.users_collection
managers_collection = config.managers_collection
camera_collection = config.camera_collection
data_collection = config.data_collection
save_path = config.save_path
greeted_employees = {}


# ----------------------------------------------------------------
def play_greeting(name, greeting_type="chào"):
    """Phát âm thanh chào mừng hoặc tạm biệt"""
    text = f"{greeting_type.capitalize()} {name}!"
    tts = gTTS(text, lang="vi")
    audio_file = "greeting.mp3"
    tts.save(audio_file)

    # Phát âm thanh bằng FFmpeg
    os.system(f"ffplay -nodisp -autoexit {audio_file}")


# ----------------------------------------------------------------
def generate_all_user_embeddings():
    """
    Cập nhật toàn bộ face_embeddings trong mongoDB, nếu thư mục hình ảnh của các users bị thay đổi
    """
    users = users_collection.find()

    for user in users:
        user_id = user["_id"]
        photo_folder = user.get("photo_folder")

        LOGGER.info(f"Processing user ID {user_id}...")

        if not photo_folder or not os.path.exists(photo_folder):
            LOGGER.warning(f"Photo folder does not exist for user ID {user_id}: {photo_folder}")
            continue

        face_embeddings = []
        photo_count = 0

        for file_name in os.listdir(photo_folder):
            file_path = os.path.join(photo_folder, file_name)

            if os.path.isfile(file_path):
                try:
                    # Gọi hàm xử lý để lấy face embedding
                    face_embedding = process_image(file_path, detector)

                    if face_embedding is not None:
                        # Lưu theo dạng danh sách {photo_name, embedding}
                        face_embeddings.append({
                            "photo_name": file_name,
                            "embedding": face_embedding.tolist()
                        })

                        photo_count += 1
                        LOGGER.info(f"Added embedding for file {file_name} (user ID {user_id})")

                except Exception as e:
                    LOGGER.error(f"Error processing file {file_path}: {e}")

        # Cập nhật face_embeddings vào users_collection
        users_collection.update_one(
            {"_id": user_id},
            {"$set": {"face_embeddings": face_embeddings}}
        )

        LOGGER.info(f"Completed processing user ID {user_id}. Total photos processed: {photo_count}")


# ----------------------------------------------------------------
def remove_accents(input_str):
    """
    Loại bỏ dấu tiếng Việt và chuyển Đ -> D, đ -> d
    """
    input_str = input_str.replace("Đ", "D").replace("đ", "d")  # Chuyển Đ -> D, đ -> d
    nfkd_form = unicodedata.normalize('NFKD', input_str)  # Chuẩn hóa Unicode
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])


# ----------------------------------------------------------------
def shorten_name(full_name):
    """
    Rút gọn tên thành dạng: N.N.Quyet (bỏ dấu và rút gọn)
    """
    full_name = remove_accents(full_name)  # Bỏ dấu trước khi rút gọn
    words = full_name.split()  # Tách tên theo khoảng trắng
    if len(words) >= 2:
        short_name = ".".join([w[0] for w in words[:-1]]) + "." + words[-1]
    else:
        short_name = words[0]  # Nếu chỉ có một từ, giữ nguyên
    return short_name


# ----------------------------------------------------------------
def build_faiss_index():
    """
    Tạo FAISS index (.faiss) và tệp ánh xạ ID (.pkl) từ MongoDB
    """
    embeddings = []
    id_mapping = {}  # Lưu mapping từ FAISS index → user (_id, full_name)

    index_counter = 0  # Đếm số embeddings

    # Lấy thông tin từ MongoDB
    users = users_collection.find({}, {"_id": 1, "full_name": 1, "face_embeddings": 1})
    
    for user in users:
        user_id = user["_id"]
        full_name = user.get("full_name", "Unknown")  # Nếu không có full_name thì gán "Unknown"

        for face_entry in user.get("face_embeddings", []):  # Duyệt qua từng embedding
            embeddings.append(face_entry["embedding"])
            id_mapping[index_counter] = {
                "id": user_id,
                "full_name": full_name
            }
            index_counter += 1

    # Kiểm tra nếu không có embeddings
    if not embeddings:
        LOGGER.warning("Không có embeddings nào trong MongoDB!")
        return

    # Chuyển thành NumPy array
    embeddings = np.array(embeddings, dtype=np.float32)

    # Xác định số chiều vector
    vector_dim = embeddings.shape[1]

    # Sử dụng FAISS với Inner Product (vì vector đã chuẩn hóa)
    index = faiss.IndexFlatIP(vector_dim)

    # Thêm vector vào FAISS Index
    index.add(embeddings)

    # Lưu FAISS Index
    faiss.write_index(index, config.faiss_file)

    # Lưu id_mapping thành file .pkl
    with open(config.faiss_mapping_file, "wb") as f:
        pickle.dump(id_mapping, f)

    LOGGER.info(f"FAISS index đã được tạo và lưu vào {config.faiss_file}.")
    LOGGER.info(f"Mapping index → user đã được lưu vào {config.faiss_mapping_file}.")


# ----------------------------------------------------------------
def build_ann_index():
    """
    Tạo tệp ann và tệp ánh xạ
    """
    embeddings = []
    id_mapping = {}  # Lưu mapping từ Annoy index → user (_id, full_name)

    index_counter = 0  # Đếm số embeddings

    # Lấy thông tin _id, full_name và embeddings từ MongoDB
    users = users_collection.find({}, {"_id": 1, "full_name": 1, "face_embeddings": 1})
    
    for user in users:
        user_id = user["_id"]
        full_name = user.get("full_name", "Unknown")  # Nếu không có full_name thì gán "Unknown"

        for face_entry in user.get("face_embeddings", []):  # Duyệt qua từng embedding
            embeddings.append(face_entry["embedding"])
            id_mapping[index_counter] = {
                "id": user_id,  # Sử dụng _id thay vì user_id
                "full_name": full_name  # Lưu full_name thay vì photo_name
            }
            index_counter += 1

    # Kiểm tra nếu không có embeddings
    if not embeddings:
        LOGGER.warning("Không có embeddings nào trong MongoDB!")
        return

    # Chuyển thành NumPy array để xử lý nhanh hơn
    embeddings = np.array(embeddings, dtype=np.float32)

    # Xác định số chiều của vector embedding
    vector_dim = len(embeddings[0])

    # Khởi tạo Annoy Index
    t = AnnoyIndex(vector_dim, 'euclidean')

    # Thêm từng embedding vào Annoy Index
    for i, emb in enumerate(embeddings):
        t.add_item(i, emb)

    # Xây dựng Annoy Index với 10 cây
    t.build(10)

    # Lưu file .ann
    t.save(config.ann_file)

    # Lưu id_mapping thành file .npy
    np.save(config.mapping_file, id_mapping)

    LOGGER.info(f"Annoy index has been successfully created and saved to {config.ann_file}.")
    LOGGER.info(f"Mapping index → user has been saved to {config.mapping_file}.")


# ----------------------------------------------------------------
@app.route('/get_all_users', methods=['GET'])
def get_all_users():
    # Lấy tham số từ query
    department_id = request.args.get('department_id', None)
    without_face_embeddings = request.args.get('without_face_embeddings', '0') == '1'  # Chuyển thành bool

    # Tạo bộ lọc truy vấn MongoDB
    query = {}
    if department_id:
        query['department_id'] = department_id  # Lọc theo phòng ban

    # Truy vấn danh sách user
    users = list(users_collection.find(query))

    # Nếu without_face_embeddings=True, xóa trường face_embeddings khỏi kết quả
    if without_face_embeddings:
        for user in users:
            user.pop('face_embeddings', None)  # Xóa nếu tồn tại

    return jsonify(users)


# ----------------------------------------------------------------
@app.route('/add_user', methods=['POST'])
def add_user():
    data = request.json
    required_fields = ["full_name", "department_id"]
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing required fields"}), 400

    # Tạo ID tự động tăng dần
    last_user = users_collection.find_one(sort=[("_id", -1)])
    user_id = last_user["_id"] + 1 if last_user else 1

    # Tạo đường dẫn thư mục cho user
    folder_path = f"{save_path}/uploads/user_{user_id}"
    
    # Chuẩn bị dữ liệu để thêm vào MongoDB
    user = {
        "_id": user_id,
        "full_name": data["full_name"],
        "department_id": data["department_id"],
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "photo_folder": folder_path
    }

    try:
        # Thêm user vào MongoDB
        users_collection.insert_one(user)

        # Tạo thư mục cho user
        os.makedirs(folder_path, exist_ok=True)

        return jsonify({"message": "User added", "id": user_id, "photo_folder": folder_path}), 201
    except Exception as e:
        return jsonify({"error": f"Failed to add user: {str(e)}"}), 500
    

# ----------------------------------------------------------------
@app.route('/delete_user/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    try:
        # Lấy thông tin user từ MongoDB
        user = users_collection.find_one({"_id": int(user_id)})
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Xóa user khỏi MongoDB
        result = users_collection.delete_one({"_id": int(user_id)})
        if result.deleted_count == 0:
            return jsonify({"error": "Failed to delete user"}), 500

        # Xóa thư mục của user
        folder_path = user.get("photo_folder")
        if folder_path and os.path.exists(folder_path):
            shutil.rmtree(folder_path)  # Xóa toàn bộ thư mục và nội dung bên trong

            # Cập nhật ann index
            build_faiss_index()

        return jsonify({"message": "User and folder deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to delete user: {str(e)}"}), 500
    

# ----------------------------------------------------------------
@app.route('/upload_photo/<user_id>', methods=['POST'])
def upload_photo(user_id):
    if 'photo' not in request.files:
        return jsonify({"error": "No photo uploaded"}), 400

    photo = request.files['photo']
    filename = photo.filename

    try:
        user_id = int(user_id)
    except ValueError:
        return jsonify({"error": "Invalid user ID"}), 400

    # Lấy thông tin user
    user = users_collection.find_one({"_id": user_id})
    if not user:
        return jsonify({"error": "User not found"}), 404

    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    existing_photos = {entry["photo_name"] for entry in user.get("face_embeddings", [])}
    if filename in existing_photos:
        return jsonify({"error": "Photo already exists"}), 400

    try:
        # Tạo đường dẫn lưu ảnh
        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, filename)

        # Lưu ảnh tạm
        photo.save(file_path)

        # Xử lý ảnh → lấy embedding
        face_embedding = process_image(file_path, detector)

        if face_embedding is None:
            # Không thành công → xoá ảnh tạm
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({"error": "Failed to process image"}), 400

        # Lưu embedding vào DB
        embedding_entry = {
            "photo_name": filename,
            "embedding": face_embedding.tolist()
        }

        users_collection.update_one(
            {"_id": user_id},
            {"$push": {"face_embeddings": embedding_entry}}
        )

        build_faiss_index()

        return jsonify({
            "message": "Photo uploaded and face features saved",
            "photo_name": filename
        }), 200

    except Exception as e:
        # Nếu có lỗi khi xử lý/lưu, cũng xoá ảnh nếu đã tồn tại
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({"error": f"Failed to upload photo: {str(e)}"}), 500
    
    
# ----------------------------------------------------------------
@app.route('/delete_photo/<user_id>', methods=['DELETE'])
def delete_photo(user_id):
    data = request.get_json()

    if not data or "file_name" not in data:
        return jsonify({"error": "File name is required"}), 400

    file_name = data["file_name"]

    # Kiểm tra xem user có tồn tại không
    user = users_collection.find_one({"_id": int(user_id)})
    if not user:
        return jsonify({"error": "User not found"}), 404

    # Kiểm tra xem thư mục user có tồn tại không
    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    # Đường dẫn đầy đủ của file cần xóa
    full_path = os.path.join(folder_path, file_name)

    if not os.path.exists(full_path):
        return jsonify({"error": "File not found"}), 404

    try:
        # Xóa file ảnh
        os.remove(full_path)

        # Xóa embedding tương ứng từ MongoDB
        users_collection.update_one(
            {"_id": int(user_id)},
            {"$pull": {"face_embeddings": {"photo_name": file_name}}}  # Xóa entry có tên ảnh tương ứng
        )

        # Build ann
        build_faiss_index()

        return jsonify({"message": "Photo and embedding deleted successfully"}), 200

    except Exception as e:
        return jsonify({"error": f"Failed to delete photo: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/get_all_managers', methods=['GET'])
def get_all_managers():
    managers = list(managers_collection.find({}, {"_id": 1, "fullname": 1, "department_id": 1, "username": 1}))
    for manager in managers:
        manager["_id"] = str(manager["_id"])
    return jsonify(managers), 200


# ----------------------------------------------------------------
@app.route('/add_manager', methods=['POST'])
def add_manager():
    data = request.json

    # Kiểm tra dữ liệu đầu vào
    required_fields = ["fullname", "department_id", "username", "password"]
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing required fields"}), 400

    # Kiểm tra username đã tồn tại chưa
    existing_manager = managers_collection.find_one({"username": data["username"]})
    if existing_manager:
        return jsonify({"error": "Username already exists"}), 409  # 409: Conflict

    # Tạo ID tự động tăng dần
    last_manager = managers_collection.find_one(sort=[("_id", -1)])
    manager_id = last_manager["_id"] + 1 if last_manager else 1

    # Thêm manager vào MongoDB
    manager = {
        "_id": manager_id,
        "fullname": data["fullname"],
        "department_id": data["department_id"],
        "username": data["username"],
        "password": data["password"],
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    try:
        managers_collection.insert_one(manager)
        return jsonify({"message": "Manager added", "id": manager_id}), 201
    except Exception as e:
        return jsonify({"error": f"Failed to add manager: {str(e)}"}), 500

# ----------------------------------------------------------------
@app.route('/delete_manager/<manager_id>', methods=['DELETE'])
def delete_manager(manager_id):
    result = managers_collection.delete_one({"_id": int(manager_id)})
    if result.deleted_count == 0:
        return jsonify({"error": "Manager not found"}), 404
    return jsonify({"message": "Manager deleted"}), 200


# ----------------------------------------------------------------
@app.route('/login', methods=['POST'])
def login():
    data = request.json  # Dữ liệu từ client gửi lên
    username = data.get('username')
    password = data.get('password')

    # Tìm manager trong MongoDB
    manager = managers_collection.find_one({"username": username})
    if not manager:
        return jsonify({"success": False, "message": "Username not found"}), 404

    # So khớp mật khẩu
    if manager["password"] == password:  # Bạn có thể dùng hash để bảo mật hơn
        return jsonify({
            "success": True,
            "message": "Login successful",
            "data": {
                "fullname": manager["fullname"],
                "department_id": manager["department_id"],
                "username": manager["username"],
                "created_at": manager["created_at"]
            }
        }), 200
    else:
        return jsonify({"success": False, "message": "Incorrect password"}), 401

    
# ----------------------------------------------------------------
@app.route('/get_attendance', methods=['GET'])
def get_attendance():
    try:
        # Lấy ngày hiện tại theo định dạng yyyy-mm-dd
        today = datetime.now().strftime("%Y-%m-%d")

        # Truy vấn tất cả bản ghi có trường date = hôm nay
        records = list(data_collection.find({"date": today}))

        # Tạo danh sách kết quả
        attendance_list = []
        for record in records:
            attendance_list.append({
                "name": record.get("name", f"Unknown ({record.get('user_id')})"),
                "check_in_time": record.get("check_in_time", "N/A"),
                "check_out_time": record.get("check_out_time", "N/A"),
            })

        return jsonify(attendance_list), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/export_attendance', methods=['POST'])
def export_attendance():
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        month = data.get("month")  # Định dạng YYYY-MM

        if not month:
            return jsonify({"error": "Thiếu thông tin month"}), 400
        
        # Lấy danh sách nhân viên từ collection users
        users = list(users_collection.find())
        results = []

        # Map ngày trong tuần sang tiếng Việt
        weekday_map = {
            0: "Thứ 2",
            1: "Thứ 3",
            2: "Thứ 4",
            3: "Thứ 5",
            4: "Thứ 6",
            5: "Thứ 7",
            6: "Chủ nhật"
        }

        # Duyệt qua từng nhân viên
        for user in users:
            user_id = user["_id"]
            full_name = user["full_name"]

            # Lặp qua từng ngày trong tháng
            start_date = datetime.strptime(f"{month}-01", "%Y-%m-%d")
            end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

            current_date = start_date
            while current_date <= end_date:
                # Chỉ xét từ thứ 2 đến thứ 6
                if current_date.weekday() < 5:
                    date_str = current_date.strftime("%Y-%m-%d")
                    weekday_str = weekday_map[current_date.weekday()]

                    # Lấy dữ liệu từ mô hình mới
                    attendance_data = data_collection.find_one({
                        "date": date_str,
                        "user_id": user_id
                    })
                    
                    # Tính giờ công và ghi chú
                    check_in, check_out, total_hours, note, kpi_deduction = calculate_work_hours_new(attendance_data)

                    # Thêm vào kết quả
                    results.append({
                        "ID": user_id,
                        "Tên nhân viên": full_name,
                        "Ngày": date_str,
                        "Thứ": weekday_str,
                        "Thời gian vào": check_in,
                        "Thời gian ra": check_out,
                        "Tổng giờ công": total_hours,
                        "Trừ KPI": kpi_deduction,
                        "Ghi chú": note
                    })

                # Sang ngày tiếp theo
                current_date += timedelta(days=1)

        # Trả về chỉ dữ liệu JSON, không tạo Excel
        return jsonify({
            "message": "Data retrieved successfully", 
            "data": results
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        edited_data = data.get("data", [])
        month_str = data.get("month", "")  # YYYY-MM
        
        if not edited_data:
            return jsonify({"error": "Không có dữ liệu để xuất"}), 400
            
        # Import các thư viện cần thiết cho định dạng Excel
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        import io
        import tempfile
        
        # Tạo đường dẫn tệp Excel động
        month_number = month_str.split("-")[1] if "-" in month_str else ""
        file_suffix = f"t{int(month_number)}" if month_number.isdigit() else "export"
        
        # Tạo tệp tạm thời
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, prefix=f'attendance_{file_suffix}_') as temp:
            temp_path = temp.name
        
        # Đảm bảo thứ tự cột đúng
        column_order = [
            "ID", 
            "Tên nhân viên", 
            "Ngày", 
            "Thứ", 
            "Thời gian vào", 
            "Thời gian ra", 
            "Tổng giờ công", 
            "Trừ KPI", 
            "Ghi chú"
        ]
        
        # Tạo DataFrame và sắp xếp cột theo thứ tự đã định
        df = pd.DataFrame(edited_data)
        
        # Đảm bảo tất cả các cột tồn tại (thêm cột thiếu nếu cần)
        for col in column_order:
            if col not in df.columns:
                df[col] = ""
        
        # Sắp xếp lại các cột theo thứ tự chỉ định
        df = df[column_order]
        
        # Xuất ra Excel với định dạng đẹp
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Chấm Công')
            
            # Lấy workbook và worksheet để định dạng
            workbook = writer.book
            worksheet = writer.sheets['Chấm Công']
            
            # Tạo định dạng cho tiêu đề
            header_font = Font(name='Calibri Light', bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Tạo định dạng cho nội dung
            content_font = Font(name='Calibri Light')
            content_alignment = Alignment(horizontal='center', vertical='center')
            
            # Áp dụng định dạng cho tiêu đề
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Áp dụng định dạng cho nội dung
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = content_font
                    cell.alignment = content_alignment
            
            # Tự động điều chỉnh chiều rộng cột
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Tạo đường viền cho bảng
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            # Áp dụng đường viền cho tất cả các ô có dữ liệu
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                        min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
        
        # Trả về đường dẫn đến tệp tạm thời để tải xuống
        return jsonify({
            "message": "Excel generated successfully", 
            "file": temp_path
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/download', methods=['GET'])
def download_file():
    file_path = request.args.get('file')
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({"error": "File không tồn tại"}), 404
    
    # Lấy tên tệp từ đường dẫn
    filename = os.path.basename(file_path)
    
    # Gửi tệp về client
    response = send_file(file_path, as_attachment=True, download_name=filename)
    
    # Xóa tệp tạm thời sau khi gửi
    @response.call_on_close
    def cleanup():
        try:
            os.remove(file_path)
        except:
            pass
    
    return response


# ----------------------------------------------------------------
def calculate_work_hours_new(attendance_data):
    """Tính tổng giờ công trong ngày, ghi chú và trừ KPI theo cấu trúc dữ liệu mới."""
    if not attendance_data:  # Không có dữ liệu
        return None, None, None, "Nghỉ", ""
    
    check_in_time = attendance_data.get("check_in_time")
    check_out_time = attendance_data.get("check_out_time")
    
    if not check_in_time or not check_out_time:
        return check_in_time, check_out_time, None, "Nghỉ", ""
    
    # Chuyển đổi định dạng thời gian
    check_in_dt = datetime.strptime(check_in_time, "%H:%M:%S")
    check_out_dt = datetime.strptime(check_out_time, "%H:%M:%S")
    
    # Tính tổng giờ công (để đơn giản, giả sử cùng ngày)
    total_hours = round((check_out_dt - check_in_dt).total_seconds() / 3600, 2)
    
    # Thời gian chuẩn để so sánh
    start_time = datetime.strptime("08:00:00", "%H:%M:%S")
    late_threshold = datetime.strptime("08:10:00", "%H:%M:%S")
    noon_deadline = datetime.strptime("11:30:00", "%H:%M:%S")
    lunch_end = datetime.strptime("12:30:00", "%H:%M:%S")
    end_time = datetime.strptime("17:30:00", "%H:%M:%S")
    
    # Kiểm tra các điều kiện
    is_late = check_in_dt > late_threshold
    is_early_leave = check_out_dt < end_time
    missing_morning = check_in_dt > noon_deadline
    missing_afternoon = check_out_dt < lunch_end
    
    # Xác định ghi chú
    note = ""
    if missing_morning:
        note = "Thiếu công sáng"
    elif missing_afternoon:
        note = "Thiếu công chiều"
    elif is_late and is_early_leave:
        note = "Đi muộn & về sớm"
    elif is_late:
        note = "Đi muộn"
    elif is_early_leave:
        note = "Về sớm"
    
    # Xác định trừ KPI - chỉ trừ khi đi muộn quá 10 phút
    kpi_deduction = 1 if is_late else ""
    
    return check_in_time, check_out_time, total_hours, note, kpi_deduction
    

# ----------------------------------------------------------------
if config.init_database:
    LOGGER.warning("Initialize database")
    generate_all_user_embeddings()
    build_faiss_index()


# ----------------------------------------------------------------
@app.route('/rebuild_all_users_embeddings', methods=['POST'])
def rebuild_all_users_embeddings():
    try:
        # Chạy xử lý ảnh người dùng
        generate_all_user_embeddings()
        
        # Cập nhật ann index
        build_faiss_index()
        
        return jsonify({"message": "User photos processed and ann index updated"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to process and update: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/get_photos/<user_id>', methods=['GET'])
def get_photos(user_id):
    # Lấy thông tin user từ MongoDB
    user = users_collection.find_one({"_id": int(user_id)})
    if not user:
        return jsonify({"error": "User not found"}), 404

    # Lấy thư mục lưu ảnh của user
    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    # Kiểm tra nếu thư mục không tồn tại
    if not os.path.exists(folder_path):
        return jsonify({"error": "User photo folder does not exist"}), 404

    try:
        # Lấy danh sách file ảnh
        photos = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

        return jsonify({"user_id": user_id, "photos": photos}), 200

    except Exception as e:
        return jsonify({"error": f"Failed to retrieve photos: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/view_photo/<user_id>/<filename>', methods=['GET'])
def view_photo(user_id, filename):
    # Lấy thông tin user từ MongoDB
    user = users_collection.find_one({"_id": int(user_id)})
    if not user:
        return jsonify({"error": "User not found"}), 404

    # Lấy thư mục lưu ảnh của user
    folder_path = user.get("photo_folder")
    if not folder_path:
        return jsonify({"error": "User folder not found"}), 500

    # Đường dẫn ảnh
    file_path = os.path.join(folder_path, filename)

    # Kiểm tra nếu file ảnh tồn tại
    if not os.path.exists(file_path):
        return jsonify({"error": "Photo not found"}), 404

    try:
        # Tự động xác định loại ảnh (PNG, JPG, JPEG, GIF, ...)
        mimetype = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'

        return send_file(file_path, mimetype=mimetype)  # Gửi ảnh với mimetype phù hợp
    except Exception as e:
        return jsonify({"error": f"Failed to load photo: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/get_user_data', methods=['GET'])
def get_user_data():
    try:
        # Lấy tham số từ request
        user_id = request.args.get("user_id", type=int)
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        camera_name = request.args.get("camera_name")

        # Kiểm tra tham số bắt buộc
        if not user_id or not start_date or not end_date or not camera_name:
            return jsonify({"error": "Missing required parameters"}), 400

        # Chuyển đổi định dạng thời gian
        try:
            start_datetime = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
            end_datetime = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
            start_date_str = start_datetime.strftime("%Y-%m-%d")
            end_date_str = end_datetime.strftime("%Y-%m-%d")
            start_time_str = start_datetime.strftime("%H:%M:%S")
            end_time_str = end_datetime.strftime("%H:%M:%S")
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD HH:MM:SS"}), 400

        # Truy vấn dữ liệu từ MongoDB
        query = {
            "user_id": user_id,
            "date": {"$gte": start_date_str, "$lte": end_date_str},
            "timestamps": {
                "$elemMatch": {
                    "camera": camera_name,
                    "time": {"$gte": start_time_str, "$lte": end_time_str}
                }
            }
        }
        user_data = list(data_collection.find(
            query,
            {"_id": 0}  # Bỏ trường _id trong kết quả trả về
        ))

        if not user_data:
            return jsonify({"error": "No data found for given criteria"}), 404

        # Tùy chỉnh kết quả trả về (lọc timestamps nếu cần)
        filtered_data = []
        for record in user_data:
            # Chỉ giữ các timestamps khớp với camera_name và khoảng thời gian
            filtered_timestamps = [
                ts for ts in record["timestamps"]
                if ts["camera"] == camera_name
                and start_time_str <= ts["time"] <= end_time_str
            ]
            if filtered_timestamps:  # Chỉ thêm bản ghi nếu có timestamps khớp
                filtered_record = record.copy()
                filtered_record["timestamps"] = filtered_timestamps
                filtered_data.append(filtered_record)

        if not filtered_data:
            return jsonify({"error": "No matching timestamps found for given criteria"}), 404

        return jsonify(filtered_data), 200

    except Exception as e:
        return jsonify({"error": f"Failed to fetch user data: {str(e)}"}), 500


# ----------------------------------------------------------------
@app.route('/get_all_cameras', methods=['GET'])
def get_all_cameras():
    cameras = list(camera_collection.find({}))
    return jsonify(cameras)


# ----------------------------------------------------------------
@app.route('/get_camera', methods=['GET'])
def get_camera():
    _id = request.args.get('_id')
    location = request.args.get('camera_location')

    if _id and location:
        return jsonify({"error": "Only filtering by _id or camera_location is allowed, not both."}), 400
    
    if _id:
        try:
            _id = int(_id)
        except ValueError:
            return jsonify({"error": "_id must be a number."}), 400
        camera = camera_collection.find_one({'_id': _id})
        if camera:
            return jsonify(camera)
        else:
            return jsonify({"error": "No camera found with the given _id."}), 404

    elif location:
        cameras = list(camera_collection.find({'camera_location': location}))
        if cameras:
            return jsonify(cameras)
        else:
            return jsonify({"error": "No cameras found with the given camera_location."}), 404
    
    else:
        return jsonify({"error": "Please provide either _id or camera_location for filtering."}), 400
    

# ----------------------------------------------------------------
@app.route('/add_camera', methods=['POST'])
def add_camera():
    data = request.get_json()

    mac_address = data.get("MAC_address")
    username = data.get("user")
    password = data.get("password")
    auto_discover = data.get("auto_discover", False)

    if not mac_address or not username or not password:
        return jsonify({"error": "MAC address, username, and password are required."}), 400

    if auto_discover is True:
        # Nếu auto_discover = True, không cho phép nhập IP và RTSP thủ công
        if "IP" in data or "RTSP" in data:
            return jsonify({"error": "When auto_discover is True, do not provide IP and RTSP. They will be generated automatically."}), 400

        result = onvif_camera_tools.find_ip_and_rtsp_by_mac(mac_address, username, password)
        if not result:
            return jsonify({"error": f"Could not find device with MAC address {mac_address}."}), 404
        ip_address = result["IP"]
        rtsp_url = result["RTSP"]
    else:
        # Người dùng bắt buộc phải nhập IP và RTSP
        ip_address = data.get("IP")
        rtsp_url = data.get("RTSP")
        if not ip_address or not rtsp_url:
            return jsonify({"error": "When auto_discover is False, IP and RTSP must be provided manually."}), 400

    # Auto increment _id
    last_camera = camera_collection.find_one(sort=[("_id", -1)])
    new_id = last_camera['_id'] + 1 if last_camera else 1

    camera = {
        "_id": new_id,
        "camera_name": data.get("camera_name"),
        "camera_location": data.get("camera_location"),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": username,
        "password": password,
        "MAC_address": mac_address,
        "IP": ip_address,
        "RTSP": rtsp_url
    }

    camera_collection.insert_one(camera)
    return jsonify({"message": "Camera added successfully.", "camera": camera}), 201
    

# ----------------------------------------------------------------
@app.route('/delete_camera', methods=['DELETE'])
def delete_camera():
    _id = request.args.get('_id')

    if not _id:
        return jsonify({"error": "_id is required for deletion."}), 400

    try:
        _id = int(_id)
    except ValueError:
        return jsonify({"error": "_id must be a number."}), 400

    result = camera_collection.delete_one({'_id': _id})

    if result.deleted_count == 0:
        return jsonify({"error": "No camera found with the given _id."}), 404

    return jsonify({"message": "Camera deleted successfully."}), 200


# ----------------------------------------------------------------
@app.route('/update_camera', methods=['PUT'])
def update_camera():
    _id = request.args.get('_id')
    data = request.get_json()

    if not _id:
        return jsonify({"error": "_id is required for updating."}), 400

    try:
        _id = int(_id)
    except ValueError:
        return jsonify({"error": "_id must be a number."}), 400

    update_fields = {key: data[key] for key in data if key != 'created_at'}

    if not update_fields:
        return jsonify({"error": "No valid fields provided for update."}), 400

    result = camera_collection.update_one({'_id': _id}, {'$set': update_fields})

    if result.matched_count == 0:
        return jsonify({"error": "No camera found with the given _id."}), 404

    return jsonify({"message": "Camera updated successfully."}), 200


# ----------------------------------------------------------------
@app.route('/update_user/<user_id>', methods=['PUT'])
def update_user(user_id):
    data = request.get_json()

    # Only allow updating these fields
    update_fields = {}
    if 'full_name' in data:
        update_fields['full_name'] = data['full_name']
    if 'department_id' in data:
        update_fields['department_id'] = data['department_id']

    if not update_fields:
        return jsonify({"error": "No valid fields to update"}), 400

    result = users_collection.update_one(
        {"_id": int(user_id)},
        {"$set": update_fields}
    )

    if result.matched_count == 0:
        return jsonify({"error": "User not found"}), 404

    return jsonify({
        "message": "User updated successfully",
        "updated_fields": update_fields
    }), 200


# ----------------------------------------------------------------
@app.route("/get_qr_code", methods=["GET"])
def get_qr_code():
    try:
        marker_id = int(request.args.get("id", 0))
        size = int(request.args.get("size", 400))
        dictionary_name = request.args.get("marker", "DICT_5X5_100")

        marker_image = generate_aruco_marker(dictionary_name, marker_id, size)

        # Encode ảnh thành PNG và đưa vào bộ nhớ RAM
        success, encoded_image = cv2.imencode(".png", marker_image)
        if not success:
            raise RuntimeError("Failed to encode marker image.")
        
        # Dùng BytesIO giả lập file
        image_io = io.BytesIO(encoded_image.tobytes())
        image_io.seek(0)

        return send_file(image_io, mimetype="image/png", as_attachment=False)

    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ---------------------------------------------------------------- 
@app.route("/users")
def users():
    return render_template("users.html")


# ----------------------------------------------------------------
if __name__ == "__main__":
    app.run(port=6123)