from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from config import config
from datetime import datetime, timedelta
import os

# Kết nối tới MongoDB
users_collection = config.users_collection
managers_collection = config.managers_collection
camera_collection = config.camera_collection
data_collection = config.data_collection
save_path = config.save_path
greeted_employees = {}


app = Flask(__name__)

"""
API
https://3hinc.nnq962.pro/get_attendance
https://3hinc.nnq962.pro/get_all_users?without_face_embeddings=1
"""


# ----------------------------------------------------------------
def calculate_work_hours_new(attendance_data):
    """Tính tổng giờ công trong ngày, ghi chú và trừ KPI theo cấu trúc dữ liệu mới."""
    if not attendance_data:  # Không có dữ liệu
        return None, None, None, "Nghỉ", ""
    
    check_in_time = attendance_data.get("check_in_time")
    check_out_time = attendance_data.get("check_out_time")
    
    if not check_in_time or not check_out_time:
        return check_in_time, check_out_time, None, "Nghỉ", ""
    
    # Chuyển đổi định dạng thời gian
    check_in_dt = datetime.strptime(check_in_time, "%H:%M:%S")
    check_out_dt = datetime.strptime(check_out_time, "%H:%M:%S")
    
    # Tính tổng giờ công (để đơn giản, giả sử cùng ngày)
    total_hours = round((check_out_dt - check_in_dt).total_seconds() / 3600, 2)
    
    # Thời gian chuẩn để so sánh
    start_time = datetime.strptime("08:00:00", "%H:%M:%S")
    late_threshold = datetime.strptime("08:10:00", "%H:%M:%S")
    noon_deadline = datetime.strptime("11:30:00", "%H:%M:%S")
    lunch_end = datetime.strptime("12:30:00", "%H:%M:%S")
    end_time = datetime.strptime("17:30:00", "%H:%M:%S")
    
    # Kiểm tra các điều kiện
    is_late = check_in_dt > late_threshold
    is_early_leave = check_out_dt < end_time
    missing_morning = check_in_dt > noon_deadline
    missing_afternoon = check_out_dt < lunch_end
    
    # Xác định ghi chú
    note = ""
    if missing_morning:
        note = "Thiếu công sáng"
    elif missing_afternoon:
        note = "Thiếu công chiều"
    elif is_late and is_early_leave:
        note = "Đi muộn & về sớm"
    elif is_late:
        note = "Đi muộn"
    elif is_early_leave:
        note = "Về sớm"
    
    # Xác định trừ KPI - chỉ trừ khi đi muộn quá 10 phút
    kpi_deduction = 1 if is_late else ""
    
    return check_in_time, check_out_time, total_hours, note, kpi_deduction


# ----------------------------------------------------------------
@app.route('/export_attendance', methods=['POST'])
def export_attendance():
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        month = data.get("month")  # Định dạng YYYY-MM

        if not month:
            return jsonify({"error": "Thiếu thông tin month"}), 400
        
        # Lấy danh sách nhân viên từ collection users
        users = list(users_collection.find())
        results = []

        # Map ngày trong tuần sang tiếng Việt
        weekday_map = {
            0: "Thứ 2",
            1: "Thứ 3",
            2: "Thứ 4",
            3: "Thứ 5",
            4: "Thứ 6",
            5: "Thứ 7",
            6: "Chủ nhật"
        }

        # Duyệt qua từng nhân viên
        for user in users:
            user_id = user["_id"]
            full_name = user["full_name"]

            # Lặp qua từng ngày trong tháng
            start_date = datetime.strptime(f"{month}-01", "%Y-%m-%d")
            end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

            current_date = start_date
            while current_date <= end_date:
                # Chỉ xét từ thứ 2 đến thứ 6
                if current_date.weekday() < 5:
                    date_str = current_date.strftime("%Y-%m-%d")
                    weekday_str = weekday_map[current_date.weekday()]

                    # Lấy dữ liệu từ mô hình mới
                    attendance_data = data_collection.find_one({
                        "date": date_str,
                        "user_id": user_id
                    })
                    
                    # Tính giờ công và ghi chú
                    check_in, check_out, total_hours, note, kpi_deduction = calculate_work_hours_new(attendance_data)

                    # Thêm vào kết quả
                    results.append({
                        "ID": user_id,
                        "Tên nhân viên": full_name,
                        "Ngày": date_str,
                        "Thứ": weekday_str,
                        "Thời gian vào": check_in,
                        "Thời gian ra": check_out,
                        "Tổng giờ công": total_hours,
                        "Trừ KPI": kpi_deduction,
                        "Ghi chú": note
                    })

                # Sang ngày tiếp theo
                current_date += timedelta(days=1)

        # Trả về chỉ dữ liệu JSON, không tạo Excel
        return jsonify({
            "message": "Data retrieved successfully", 
            "data": results
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        edited_data = data.get("data", [])
        month_str = data.get("month", "")  # YYYY-MM
        
        if not edited_data:
            return jsonify({"error": "Không có dữ liệu để xuất"}), 400
            
        # Import các thư viện cần thiết cho định dạng Excel
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        import io
        import tempfile
        
        # Tạo đường dẫn tệp Excel động
        month_number = month_str.split("-")[1] if "-" in month_str else ""
        file_suffix = f"t{int(month_number)}" if month_number.isdigit() else "export"
        
        # Tạo tệp tạm thời
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, prefix=f'attendance_{file_suffix}_') as temp:
            temp_path = temp.name
        
        # Đảm bảo thứ tự cột đúng
        column_order = [
            "ID", 
            "Tên nhân viên", 
            "Ngày", 
            "Thứ", 
            "Thời gian vào", 
            "Thời gian ra", 
            "Tổng giờ công", 
            "Trừ KPI", 
            "Ghi chú"
        ]
        
        # Tạo DataFrame và sắp xếp cột theo thứ tự đã định
        df = pd.DataFrame(edited_data)
        
        # Đảm bảo tất cả các cột tồn tại (thêm cột thiếu nếu cần)
        for col in column_order:
            if col not in df.columns:
                df[col] = ""
        
        # Sắp xếp lại các cột theo thứ tự chỉ định
        df = df[column_order]
        
        # Xuất ra Excel với định dạng đẹp
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Chấm Công')
            
            # Lấy workbook và worksheet để định dạng
            workbook = writer.book
            worksheet = writer.sheets['Chấm Công']
            
            # Tạo định dạng cho tiêu đề
            header_font = Font(name='Calibri Light', bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Tạo định dạng cho nội dung
            content_font = Font(name='Calibri Light')
            content_alignment = Alignment(horizontal='center', vertical='center')
            
            # Áp dụng định dạng cho tiêu đề
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Áp dụng định dạng cho nội dung
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = content_font
                    cell.alignment = content_alignment
            
            # Tự động điều chỉnh chiều rộng cột
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Tạo đường viền cho bảng
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            # Áp dụng đường viền cho tất cả các ô có dữ liệu
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                        min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
        
        # Trả về đường dẫn đến tệp tạm thời để tải xuống
        return jsonify({
            "message": "Excel generated successfully", 
            "file": temp_path
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route('/download', methods=['GET'])
def download_file():
    file_path = request.args.get('file')
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({"error": "File không tồn tại"}), 404
    
    # Lấy tên tệp từ đường dẫn
    filename = os.path.basename(file_path)
    
    # Gửi tệp về client
    response = send_file(file_path, as_attachment=True, download_name=filename)
    
    # Xóa tệp tạm thời sau khi gửi
    @response.call_on_close
    def cleanup():
        try:
            os.remove(file_path)
        except:
            pass
    
    return response


# ----------------------------------------------------------------
@app.route('/get_all_users', methods=['GET'])
def get_all_users():
    # Lấy tham số từ query
    department_id = request.args.get('department_id', None)
    without_face_embeddings = request.args.get('without_face_embeddings', '0') == '1'  # Chuyển thành bool

    # Tạo bộ lọc truy vấn MongoDB
    query = {}
    if department_id:
        query['department_id'] = department_id  # Lọc theo phòng ban

    # Truy vấn danh sách user
    users = list(users_collection.find(query))

    # Nếu without_face_embeddings=True, xóa trường face_embeddings khỏi kết quả
    if without_face_embeddings:
        for user in users:
            user.pop('face_embeddings', None)  # Xóa nếu tồn tại

    return jsonify(users)


# ----------------------------------------------------------------
@app.route('/get_attendance', methods=['GET'])
def get_attendance():
    try:
        # Lấy ngày hiện tại theo định dạng yyyy-mm-dd
        today = datetime.now().strftime("%Y-%m-%d")

        # Truy vấn tất cả bản ghi có trường date = hôm nay
        records = list(data_collection.find({"date": today}))

        # Tạo danh sách kết quả
        attendance_list = []
        for record in records:
            attendance_list.append({
                "name": record.get("name", f"Unknown ({record.get('user_id')})"),
                "check_in_time": record.get("check_in_time", "N/A"),
                "check_out_time": record.get("check_out_time", "N/A"),
            })

        return jsonify(attendance_list), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ----------------------------------------------------------------
@app.route("/")
def home():
    return render_template("view.html")


# ----------------------------------------------------------------
@app.route("/export")
def export():
    return render_template("export.html")


# ----------------------------------------------------------------
if __name__ == "__main__":
    app.run(port=5555)
