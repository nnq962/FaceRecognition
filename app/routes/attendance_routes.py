from flask import Blueprint, request, jsonify, send_file, session
from datetime import datetime, timedelta, date
import tempfile
import pandas as pd
import os
from bson.objectid import ObjectId

import pymongo
from config import config
from app.auth import login_required, role_required
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import mimetypes
import uuid
from werkzeug.utils import secure_filename
from app import user_collection, reports_collection, auth
from utils.logger_config import LOGGER

# Khởi tạo Blueprint
attendance_bp = Blueprint('attendance_routes', __name__)
all_room_logs = config.database['all_room_logs']

# Cấu hình cho uploads
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx'}
machine_error_report_type = {
    "no_recognize": "Máy không nhận diện",
    "wrong_time": "Máy ghi sai giờ",
    "device_off": "Máy không hoạt động",
    "other": "Lỗi khác"
}

leave_request_report_type = {
    'late': 'Xin phép đi muộn',
    'early_leave': 'Xin phép về sớm',
    'absent': 'Xin phép nghỉ',
    'other': 'Khác'
}

# ------------------------------------------------------------
@attendance_bp.route('/get_reports', methods=['GET'])
def get_reports():
    """
    API lấy danh sách báo cáo theo phân quyền
    
    Phân quyền:
    - Admin: Chỉ thấy reports của user cùng room trong ngày hiện tại
    - Super_admin: Thấy tất cả reports trong ngày hiện tại  
    - Super_admin + username "nnq962": Thấy tất cả reports từ đầu tháng đến hiện tại
    
    Query parameters:
    - report_type: Lọc theo loại báo cáo (optional): all, incorrect_photo, machine_error, leave_request
    - status: Lọc theo trạng thái (optional): all, pending, approved, rejected
    - from_date: Từ ngày (optional, format: YYYY-MM-DD)
    - to_date: Đến ngày (optional, format: YYYY-MM-DD)
    
    Returns:
        JSON response với danh sách reports và thông tin user
    """
    try:
        # Lấy thông tin user từ session
        session_token = session.get('session_token')
        user = auth.get_user_by_session_token(session_token)
        
        if not user:
            return jsonify({
                "status": "error",
                "message": "Unauthorized"
            }), 401
            
        role = user.get("role")
        username = user.get("username")
        user_room_id = user.get('room_id')
        
        LOGGER.debug(f"Role: {role}")
        LOGGER.debug(f"Username: {username}")
        LOGGER.debug(f"Room id: {user_room_id}")
        
        # Kiểm tra quyền truy cập
        if role not in ['admin', 'super_admin']:
            return jsonify({
                "status": "error",
                "message": "Access denied. Admin or Super Admin role required."
            }), 403
        
        # Xác định khoảng thời gian query
        now = datetime.now()
        
        # Trường hợp đặc biệt: super_admin + username "nnq962" -> lấy từ đầu tháng
        if role == 'super_admin' and username == 'nnq962':
            start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            default_from_date = start_of_month
            default_to_date = end_of_day
        else:
            # Mặc định: chỉ lấy ngày hiện tại
            start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            default_from_date = start_of_day
            default_to_date = end_of_day
        
        # Lấy parameters từ query string
        report_type_filter = request.args.get('report_type', 'all')  # THÊM: Lấy report_type filter
        status_filter = request.args.get('status', 'all')
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')

        LOGGER.debug(f"From date: {from_date_str}")
        LOGGER.debug(f"To date: {to_date_str}")
        
        # THÊM: Debug log để kiểm tra parameters
        LOGGER.debug(f"Report type filter: {report_type_filter}")
        LOGGER.debug(f"Status filter: {status_filter}")
        
        # Xử lý date parameters
        if from_date_str:
            try:
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            except ValueError:
                return jsonify({
                    "status": "error",
                    "message": "Invalid from_date format. Use YYYY-MM-DD"
                }), 400
        else:
            from_date = default_from_date
            
        if to_date_str:
            try:
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            except ValueError:
                return jsonify({
                    "status": "error",
                    "message": "Invalid to_date format. Use YYYY-MM-DD"
                }), 400
        else:
            to_date = default_to_date
        
        # Xây dựng query filter cho reports
        reports_filter = {
            "created_at": {
                "$gte": from_date,
                "$lte": to_date
            }
        }
        
        # THÊM: Lọc theo report_type nếu không phải "all"
        if report_type_filter != 'all':
            reports_filter["report_type"] = report_type_filter
            LOGGER.debug(f"Added report_type filter: {report_type_filter}")
        
        # Lọc theo status nếu không phải "all"
        if status_filter != 'all':
            reports_filter["status"] = status_filter
        
        # THÊM: Debug log để kiểm tra query filter cuối cùng
        LOGGER.debug(f"Final reports filter: {reports_filter}")
        
        # Xác định user_ids được phép xem dựa trên role
        if role == 'admin':
            # Admin chỉ thấy reports của users cùng room
            users_in_same_room = list(user_collection.find(
                {"room_id": user_room_id, "active": True},
                {"user_id": 1}
            ))
            allowed_user_ids = [u["user_id"] for u in users_in_same_room]
            
            if not allowed_user_ids:
                return jsonify({
                    "status": "success",
                    "data": {
                        "reports": [],
                        "total_count": 0,
                        "summary": {
                            "pending": 0,
                            "approved": 0,
                            "rejected": 0
                        }
                    }
                }), 200
                
            reports_filter["user_id"] = {"$in": allowed_user_ids}
            
        elif role == 'super_admin':
            # Super admin thấy tất cả reports (không cần thêm filter user_id)
            pass
        
        # Lấy reports từ database
        reports_cursor = reports_collection.find(reports_filter).sort("created_at", -1)
        reports = list(reports_cursor)
        
        # THÊM: Debug log để kiểm tra số lượng reports tìm được
        LOGGER.debug(f"Found {len(reports)} reports matching filter")
        
        # Lấy thông tin users liên quan
        user_ids = list(set([report["user_id"] for report in reports]))
        users_info = {}
        
        if user_ids:
            users_cursor = user_collection.find(
                {"user_id": {"$in": user_ids}},
                {"user_id": 1, "name": 1, "room_id": 1}
            )
            for user_info in users_cursor:
                users_info[user_info["user_id"]] = {
                    "name": user_info.get("name", "Unknown"),
                    "room_id": user_info.get("room_id", "")
                }
        
        # Xử lý và format dữ liệu trả về
        formatted_reports = []
        summary = {"pending": 0, "approved": 0, "rejected": 0}
        
        for report in reports:
            # Đếm summary
            status = report.get("status", "pending")
            if status in summary:
                summary[status] += 1
            
            # Format report data
            user_info = users_info.get(report["user_id"], {"name": "Unknown User", "room_id": ""})
            
            formatted_report = {
                "report_id": str(report["_id"]),
                "report_type": report["report_type"],
                "user_id": report["user_id"],
                "user_name": user_info["name"],
                "user_room_id": user_info["room_id"],
                "date": report["date"],
                "status": report["status"],
                "description": report.get("description", ""),
                "admin_note": report.get("admin_note", ""),
                "created_at": report["created_at"].isoformat() if isinstance(report["created_at"], datetime) else str(report["created_at"]),
                "updated_at": report["updated_at"].isoformat() if isinstance(report["updated_at"], datetime) else str(report["updated_at"]),
                "updated_by": report.get("updated_by", "")  # THÊM: Trường updated_by
            }
            
            # Thêm fields đặc biệt theo từng loại report
            if report["report_type"] == "incorrect_photo":
                formatted_report["photo_type"] = report.get("photo_type", "")
                
            elif report["report_type"] == "machine_error":
                formatted_report["error_type"] = report.get("error_type", "")
                formatted_report["error_time"] = report.get("error_time", "")
                
            elif report["report_type"] == "leave_request":
                formatted_report["request_type"] = report.get("request_type", "")
                formatted_report["attached_files"] = report.get("attached_files", [])
                
                # Format end_date nếu có (để hiển thị khoảng thời gian)
                if "end_date" in report:
                    formatted_report["end_date"] = report["end_date"]
            
            formatted_reports.append(formatted_report)
        
        return jsonify({
            "status": "success",
            "data": {
                "reports": formatted_reports,
                "total_count": len(formatted_reports),
                "summary": summary,
                "filter_info": {
                    "role": role,
                    "username": username,
                    "room_id": user_room_id if role == 'admin' else None,
                    "from_date": from_date.strftime('%Y-%m-%d'),
                    "to_date": to_date.strftime('%Y-%m-%d'),
                    "report_type_filter": report_type_filter,  # THÊM: Thông tin filter
                    "status_filter": status_filter
                }
            }
        }), 200
        
    except Exception as e:
        LOGGER.error(f"Error in get_reports: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Failed to fetch reports: {str(e)}"
        }), 500


# ------------------------------------------------------------
@attendance_bp.route('/update_report_status/<report_id>', methods=['POST'])
def update_report_status(report_id):
    """
    API cập nhật trạng thái báo cáo (phê duyệt/từ chối)
    
    Body: JSON
    {
        "status": "approved" hoặc "rejected",
        "admin_note": "Phản hồi từ quản lý (optional)"
    }
    
    Đặc biệt: Nếu approve report machine_error có error_time, sẽ tự động update attendance
    
    Returns:
        JSON response với kết quả cập nhật
    """
    try:
        # Lấy thông tin user từ session
        session_token = session.get('session_token')
        user = auth.get_user_by_session_token(session_token)
        
        if not user:
            return jsonify({
                "status": "error",
                "message": "Unauthorized"
            }), 401
            
        role = user.get("role")
        if role not in ['admin', 'super_admin']:
            return jsonify({
                "status": "error",
                "message": "Access denied. Admin or Super Admin role required."
            }), 403
        
        # Validate request data
        data = request.json
        if not data:
            return jsonify({
                "status": "error",
                "message": "Invalid JSON data"
            }), 400
            
        new_status = data.get("status")
        admin_note = data.get("admin_note", "")
        
        if new_status not in ["approved", "rejected"]:
            return jsonify({
                "status": "error",
                "message": "Invalid status. Must be 'approved' or 'rejected'"
            }), 400
        
        # Kiểm tra report có tồn tại không
        try:
            report = reports_collection.find_one({"_id": ObjectId(report_id)})
        except:
            return jsonify({
                "status": "error",
                "message": "Invalid report ID"
            }), 400
            
        if not report:
            return jsonify({
                "status": "error",
                "message": "Report not found"
            }), 404
        
        # Kiểm tra quyền cập nhật (admin chỉ được update reports của users cùng room)
        if role == 'admin':
            user_room_id = user.get('room_id')
            report_user = user_collection.find_one({"user_id": report["user_id"]})
            
            if not report_user or report_user.get("room_id") != user_room_id:
                return jsonify({
                    "status": "error",
                    "message": "Access denied. You can only update reports from your room."
                }), 403
        
        # THÊM: Xử lý đặc biệt cho machine_error report khi approved
        attendance_update_result = None
        if (new_status == "approved" and 
            report["report_type"] == "machine_error" and 
            report.get("error_time")):
            
            try:
                attendance_update_result = update_attendance_for_machine_error(report)
                LOGGER.info(f"Attendance updated for report {report_id}: {attendance_update_result}")
            except Exception as e:
                LOGGER.error(f"Failed to update attendance for report {report_id}: {str(e)}")
                # Không return error, chỉ log để không block việc approve report
        
        # Cập nhật report với thông tin người cập nhật
        update_data = {
            "status": new_status,
            "admin_note": admin_note,
            "updated_at": datetime.now(),
            "updated_by": user.get("name")
        }
        
        result = reports_collection.update_one(
            {"_id": ObjectId(report_id)},
            {"$set": update_data}
        )
        
        if result.modified_count > 0:
            response_data = {
                "report_id": report_id,
                "new_status": new_status,
                "admin_note": admin_note,
                "updated_by": user.get("name")
            }
            
            # Thêm thông tin attendance update nếu có
            if attendance_update_result:
                response_data["attendance_update"] = attendance_update_result
            
            return jsonify({
                "status": "success",
                "message": f"Report {new_status} successfully",
                "data": response_data
            }), 200
        else:
            return jsonify({
                "status": "error",
                "message": "Failed to update report"
            }), 500
            
    except Exception as e:
        LOGGER.error(f"Error in update_report_status: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Failed to update report: {str(e)}"
        }), 500


# ------------------------------------------------------------
@attendance_bp.route('/get_pending_reports_count', methods=['GET'])
def get_pending_reports_count():
    """
    API lấy số lượng báo cáo cần phê duyệt
    
    Phân quyền giống get_reports:
    - Admin: Chỉ thấy reports của user cùng room trong ngày hiện tại
    - Super_admin: Thấy tất cả reports trong ngày hiện tại  
    - Super_admin + username "nnq962": Thấy tất cả reports từ đầu tháng đến hiện tại
    
    Query parameters:
    - date: Ngày cụ thể (optional, format: YYYY-MM-DD)
    - report_type: Loại báo cáo (optional): all, incorrect_photo, machine_error, leave_request
    """
    try:
        # Lấy thông tin user từ session
        session_token = session.get('session_token')
        user = auth.get_user_by_session_token(session_token)
        
        if not user:
            return jsonify({
                "status": "error",
                "message": "Unauthorized"
            }), 401
            
        role = user.get("role")
        username = user.get("username")
        user_room_id = user.get('room_id')

        LOGGER.debug(f"Role: {role}")
        LOGGER.debug(f"Username: {username}")
        LOGGER.debug(f"Room id: {user_room_id}")

        # Kiểm tra quyền truy cập
        if role not in ['admin', 'super_admin']:
            return jsonify({
                "status": "error",
                "message": "Access denied. Admin or Super Admin role required."
            }), 403

        # Lấy parameters
        target_date = request.args.get('date')
        report_type = request.args.get('report_type', 'all')

        # Xác định khoảng thời gian query
        now = datetime.now()
        
        if target_date:
            # Nếu có target_date cụ thể, sử dụng ngày đó
            try:
                from_date = datetime.strptime(target_date, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
                to_date = datetime.strptime(target_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            except ValueError:
                return jsonify({
                    "status": "error",
                    "message": "Invalid date format. Use YYYY-MM-DD"
                }), 400
        else:
            # Trường hợp đặc biệt: super_admin + username "nnq962" -> lấy từ đầu tháng
            if role == 'super_admin' and username == 'nnq962':
                start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                from_date = start_of_month
                to_date = end_of_day
            else:
                # Mặc định: chỉ lấy ngày hiện tại
                start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                from_date = start_of_day
                to_date = end_of_day

        # Xây dựng query filter
        query = {
            "status": "pending",
            "created_at": {
                "$gte": from_date,
                "$lte": to_date
            }
        }

        # Lọc theo report_type nếu không phải "all"
        if report_type and report_type != 'all':
            valid_report_types = ['incorrect_photo', 'machine_error', 'leave_request']
            if report_type in valid_report_types:
                query['report_type'] = report_type
            else:
                return jsonify({
                    "status": "error",
                    "message": f"Invalid report_type. Valid values: {', '.join(valid_report_types + ['all'])}"
                }), 400

        # Xác định user_ids được phép xem dựa trên role (giống get_reports)
        if role == 'admin':
            # Admin chỉ thấy reports của users cùng room
            users_in_same_room = list(user_collection.find(
                {"room_id": user_room_id, "active": True},
                {"user_id": 1}
            ))
            allowed_user_ids = [u["user_id"] for u in users_in_same_room]
            
            if not allowed_user_ids:
                return jsonify({
                    "status": "success",
                    "message": "Pending report count retrieved successfully",
                    "data": {
                        "date_range": {
                            "from": from_date.strftime('%Y-%m-%d'),
                            "to": to_date.strftime('%Y-%m-%d')
                        },
                        "query_type": report_type,
                        "total_pending": 0,
                        "breakdown": {
                            'incorrect_photo': 0,
                            'machine_error': 0,
                            'leave_request': 0
                        },
                        "filter_info": {
                            "role": role,
                            "room_id": user_room_id
                        }
                    }
                }), 200
                
            query["user_id"] = {"$in": allowed_user_ids}
            
        elif role == 'super_admin':
            # Super admin thấy tất cả reports (không cần thêm filter user_id)
            pass

        LOGGER.debug(f"Final query for pending count: {query}")

        # Đếm tổng số pending reports
        pending_count = reports_collection.count_documents(query)

        # Tạo pipeline cho breakdown theo report_type
        breakdown_query = query.copy()
        # Xóa report_type filter để lấy breakdown cho tất cả types
        if 'report_type' in breakdown_query:
            del breakdown_query['report_type']

        pipeline = [
            {"$match": breakdown_query},
            {"$group": {
                "_id": "$report_type",
                "count": {"$sum": 1}
            }}
        ]
        breakdown_result = list(reports_collection.aggregate(pipeline))

        breakdown = {
            'incorrect_photo': 0,
            'machine_error': 0,
            'leave_request': 0
        }
        for item in breakdown_result:
            if item['_id'] in breakdown:
                breakdown[item['_id']] = item['count']

        LOGGER.debug(f"Pending count: {pending_count}, Breakdown: {breakdown}")

        return jsonify({
            "status": "success",
            "message": "Pending report count retrieved successfully",
            "data": {
                "date_range": {
                    "from": from_date.strftime('%Y-%m-%d'),
                    "to": to_date.strftime('%Y-%m-%d')
                },
                "query_type": report_type,
                "total_pending": pending_count,
                "breakdown": breakdown,
                "filter_info": {
                    "role": role,
                    "username": username,
                    "room_id": user_room_id if role == 'admin' else None
                }
            }
        }), 200

    except Exception as e:
        LOGGER.error(f"Error in get_pending_reports_count: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Failed to retrieve pending report count: {str(e)}"
        }), 500


# ------------------------------------------------------------
@attendance_bp.route('/download_report_file/<user_id>/<filename>', methods=['GET'])
def download_report_file(user_id, filename):
    """
    API tải file đính kèm của báo cáo
    """
    try:
        user = user_collection.find_one({"user_id": user_id})
        if not user:
            return jsonify({
                "status": "error",
                "message": "Người dùng không tồn tại"
            }), 404

        report = config.reports_collection.find_one(
            {"attached_files.saved_filename": filename, "user_id": user_id}
        )
        if not report:
            return jsonify({
                "status": "error",
                "message": "File không tồn tại hoặc bạn không có quyền truy cập"
            }), 404

        file_info = next(
            (f for f in report.get('attached_files', []) if f.get('saved_filename') == filename),
            None
        )
        if not file_info:
            return jsonify({
                "status": "error",
                "message": "Không tìm thấy thông tin file"
            }), 404

        relative_folder_path = user.get("data_folder", "")
        file_path = os.path.join(
            config.BASE_DIR,
            relative_folder_path,
            "documents",
            filename
        )
        if not os.path.exists(file_path):
            return jsonify({
                "status": "error",
                "message": "File không tồn tại trên server"
            }), 404

        mimetype = file_info.get('file_type') or mimetypes.guess_type(file_path)[0] or 'application/octet-stream'

        return send_file(
            file_path,
            mimetype=mimetype,
            as_attachment=True,
            download_name=file_info.get('original_filename', filename)
        )

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Lỗi xử lý: {str(e)}"
        }), 500


# ------------------------------------------------------------
@attendance_bp.route('/get_user_reports/<user_id>', methods=['GET'])
def get_user_reports(user_id):
    try:
        user = user_collection.find_one({"user_id": user_id})
        if not user:
            return jsonify({
                "status": "error",
                "message": "User not found"
            }), 404

        report_type = request.args.get('report_type')
        status = request.args.get('status')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        skip = (page - 1) * limit

        query = {"user_id": user_id}

        if report_type and report_type not in ['all', 'undefined', 'null']:
            query['report_type'] = report_type

        if status and status not in ['all', 'undefined', 'null']:
            query['status'] = status

        if from_date or to_date:
            date_query = {}
            if from_date and from_date not in ['undefined', 'null']:
                date_query['$gte'] = from_date
            if to_date and to_date not in ['undefined', 'null']:
                date_query['$lte'] = to_date
            if date_query:
                query['date'] = date_query

        total_count = config.reports_collection.count_documents(query)

        reports = list(config.reports_collection.find(query)
                       .sort('created_at', pymongo.DESCENDING)
                       .skip(skip)
                       .limit(limit))

        for report in reports:
            report['_id'] = str(report['_id'])
            if report.get('created_at'):
                report['created_at'] = report['created_at'].isoformat()
            if report.get('updated_at'):
                report['updated_at'] = report['updated_at'].isoformat()

        return jsonify({
            "status": "success",
            "message": "Reports retrieved successfully",
            "data": {
                "total": total_count,
                "page": page,
                "limit": limit,
                "total_pages": max(1, (total_count + limit - 1) // limit),
                "reports": reports
            }
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to retrieve reports: {str(e)}"
        }), 500


# ------------------------------------------------------------
@attendance_bp.route('/create_report/<user_id>', methods=['POST'])
def create_report(user_id):
    """
    API tạo báo cáo (dùng chung cho tất cả các loại báo cáo)
    
    Endpoint: /create_report/<user_id>
    
    Body: form-data hoặc JSON (tùy vào report_type)
    
    Form-data (cho leave_request có file đính kèm):
    - report_type: loại báo cáo (required)
    - date: Ngày báo cáo liên quan tới, format YYYY-MM-DD (required)
    - [Các trường khác tùy loại báo cáo]
    - files[]: File đính kèm (optional, cho leave_request)
    
    JSON (cho incorrect_photo và machine_error):
    {
        "report_type": "incorrect_photo", // hoặc "machine_error"
        "date": "2025-05-18",
        [Các trường khác tùy loại báo cáo]
    }
    
    Các loại báo cáo và trường dữ liệu bổ sung:
    
    1. incorrect_photo (Báo ảnh không phải tôi):
       - photo_type: "check_in" hoặc "check_out"
       - description: Lý do báo cáo
    
    2. machine_error (Báo lỗi máy chấm công):
       - error_type: Loại lỗi ("no_recognize", "wrong_time", "device_off", "other")
       - description: Mô tả chi tiết (optional)
       - error_time: Thời gian lỗi ("morning", "afternoon", "full_day") (required)
    
    3. leave_request (Gửi giấy tờ xin phép):
       - request_type: Loại giấy tờ ("late", "early_leave", "absent", "other")
       - description: Mô tả lý do (optional)
       - files[]: File đính kèm (optional)
    
    Returns:
        JSON response với kết quả báo cáo
    """
    try:
        is_form_data = request.content_type and 'multipart/form-data' in request.content_type
        data = request.form if is_form_data else request.json

        required_fields = ['report_type', 'date']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    "status": "error",
                    "message": f"Missing required field: {field}"
                }), 400

        report_type = data.get('report_type')
        valid_report_types = ['incorrect_photo', 'machine_error', 'leave_request']
        if report_type not in valid_report_types:
            return jsonify({
                "status": "error",
                "message": f"Invalid report_type. Valid values: {', '.join(valid_report_types)}"
            }), 400

        reported_user = user_collection.find_one({"user_id": user_id})
        if not reported_user:
            return jsonify({
                "status": "error",
                "message": "User not found"
            }), 404

        date = data.get('date')
        report = {
            "report_type": report_type,
            "user_id": user_id,
            "date": date,
            "status": "pending",
            "admin_note": "",
            "created_at": datetime.now(),
            "updated_at": datetime.now()
        }

        if report_type == 'incorrect_photo':
            if 'photo_type' not in data or 'description' not in data:
                return jsonify({
                    "status": "error",
                    "message": "Missing required fields for incorrect_photo"
                }), 400

            photo_type = data.get('photo_type')
            if photo_type not in ['check_in', 'check_out']:
                return jsonify({
                    "status": "error",
                    "message": "Invalid photo_type. Must be 'check_in' or 'check_out'"
                }), 400

            report["photo_type"] = photo_type
            report["description"] = data.get('description')

        elif report_type == 'machine_error':
            if 'error_type' not in data:
                return jsonify({
                    "status": "error",
                    "message": "Missing error_type"
                }), 400

            if 'error_time' not in data:
                return jsonify({
                    "status": "error",
                    "message": "Missing error_time"
                }), 400

            valid_error_types = ['no_recognize', 'wrong_time', 'device_off', 'other']
            if data.get('error_type') not in valid_error_types:
                return jsonify({
                    "status": "error",
                    "message": "Invalid error_type"
                }), 400

            # Validate error_time (now required)
            error_time = data.get('error_time')
            valid_error_times = ['morning', 'afternoon', 'full_day']
            if error_time not in valid_error_times:
                return jsonify({
                    "status": "error",
                    "message": f"Invalid error_time. Valid values: {', '.join(valid_error_times)}"
                }), 400

            report["error_type"] = data.get('error_type')
            report["description"] = data.get('description', '')
            report["error_time"] = error_time

        elif report_type == 'leave_request':
            if 'request_type' not in data:
                return jsonify({
                    "status": "error",
                    "message": "Missing request_type"
                }), 400

            valid_request_types = ['late', 'early_leave', 'absent', 'other']
            if data.get('request_type') not in valid_request_types:
                return jsonify({
                    "status": "error",
                    "message": "Invalid request_type"
                }), 400

            report["request_type"] = data.get('request_type')
            report["description"] = data.get('description', '')
            attached_files = []

            if is_form_data and 'files[]' in request.files:
                files = request.files.getlist('files[]')
                for file in files:
                    if file and allowed_file(file.filename):
                        original_filename = secure_filename(file.filename)
                        unique_filename = generate_unique_filename(original_filename)
                        relative_folder_path = reported_user.get("data_folder", "")
                        documents_folder = os.path.join(config.BASE_DIR, relative_folder_path, "documents")
                        os.makedirs(documents_folder, exist_ok=True)
                        file_path = os.path.join(documents_folder, unique_filename)
                        file.save(file_path)

                        attached_files.append({
                            "original_filename": original_filename,
                            "saved_filename": unique_filename,
                            "file_type": file.content_type or mimetypes.guess_type(original_filename)[0],
                            "file_size": os.path.getsize(file_path)
                        })

            report["attached_files"] = attached_files

        result = reports_collection.insert_one(report)

        if result.inserted_id:
            try:
                name = reported_user.get('name', 'User')
                if report_type == 'incorrect_photo':
                    title = f"Incorrect photo report from {name}"
                    content = f"{name} reported incorrect {report.get('photo_type')} photo on {date}."
                elif report_type == 'machine_error':
                    title = f"Machine error report from {name}"
                    content = f"{name} reported a machine error on {date}."
                else:
                    title = f"Leave request from {name}"
                    content = f"{name} submitted a leave request for {date}."

                # send_notification_to_admin(title, content)
            except Exception:
                pass

            return jsonify({
                "status": "success",
                "message": "Report submitted successfully",
                "data": {
                    "report_id": str(result.inserted_id)
                }
            }), 201

        else:
            return jsonify({
                "status": "error",
                "message": "Failed to save report"
            }), 500

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to process request: {str(e)}"
        }), 500  


# ------------------------------------------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ------------------------------------------------------------
# Tạo tên file an toàn và unique
def generate_unique_filename(original_filename):
    ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
    unique_name = f"{uuid.uuid4().hex}.{ext}" if ext else f"{uuid.uuid4().hex}"
    return unique_name


# ------------------------------------------------------------
@attendance_bp.route('/view_attendance_photo/<user_id>', methods=['GET'])
def view_attendance_photo(user_id):
    date_str = request.args.get("date")
    photo_type = request.args.get("type")

    if not date_str or photo_type not in ["check_in", "check_out"]:
        return jsonify({
            "status": "error",
            "message": "Missing or invalid 'date' or 'type' parameter"
        }), 400

    user = user_collection.find_one({"user_id": user_id})
    if not user:
        return jsonify({
            "status": "error",
            "message": "User not found"
        }), 404

    relative_folder_path = user.get("data_folder", "")
    date_folder_path = os.path.join(
        config.BASE_DIR,
        relative_folder_path,
        "attendance_photos",
        date_str
    )

    for ext in ['png', 'jpg', 'jpeg']:
        file_path = os.path.join(date_folder_path, f"{photo_type}.{ext}")
        if os.path.exists(file_path):
            mimetype = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
            return send_file(file_path, mimetype=mimetype)

    return jsonify({
        "status": "error",
        "message": f"{photo_type} photo for {date_str} not found"
    }), 404


# ------------------------------------------------------------
@attendance_bp.route('/get_attendance', methods=['GET'])
def get_attendance():
    try:
        user_id = request.args.get('user_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not user_id:
            return jsonify({
                "status": "error",
                "message": "Missing user_id"
            }), 400

        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    "status": "error",
                    "message": "Invalid start_date format. Use YYYY-MM-DD"
                }), 400
        else:
            today = datetime.now().date()
            start_date = date(today.year, today.month, 1)

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    "status": "error",
                    "message": "Invalid end_date format. Use YYYY-MM-DD"
                }), 400
        else:
            end_date = datetime.now().date()

        if start_date > end_date:
            return jsonify({
                "status": "error",
                "message": "Start date must be earlier than or equal to end date"
            }), 400

        attendance_records = list(config.database['all_room_logs'].find({
            'user_id': user_id,
            'date': {
                '$gte': start_date.strftime('%Y-%m-%d'),
                '$lte': end_date.strftime('%Y-%m-%d')
            }
        }).sort('date', 1))

        result = []
        for record in attendance_records:
            result.append({
                'date': record.get('date'),
                'check_in_time': record.get('check_in_time'),
                'check_out_time': record.get('check_out_time')
            })

        return jsonify({
            "status": "success",
            "message": "Attendance records retrieved successfully",
            "data": {
                "user_id": user_id,
                "start_date": start_date.strftime('%Y-%m-%d'),
                "end_date": end_date.strftime('%Y-%m-%d'),
                "attendance_records": result,
                "total_records": len(result)
            }
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to retrieve attendance: {str(e)}"
        }), 500


# ------------------------------------------------------------
@attendance_bp.route('/get_attendances', methods=['GET'])
@login_required
@role_required('get_attendances')
def get_attendances():
    try:
        # Lấy ngày hiện tại theo định dạng yyyy-mm-dd
        today = datetime.now().strftime("%Y-%m-%d")

        # Truy vấn tất cả bản ghi có trường date = hôm nay
        records = list(config.database['all_room_logs'].find({"date": today}))

        # Tạo danh sách kết quả
        attendance_list = []
        for record in records:
            attendance_list.append({
                "name": record.get("name", f"Unknown ({record.get('user_id')})"),
                "check_in_time": record.get("check_in_time", "N/A"),
                "check_out_time": record.get("check_out_time", "N/A"),
            })

        return jsonify(attendance_list), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ------------------------------------------------------------  
@attendance_bp.route('/export_attendance', methods=['POST'])
def export_attendance():
    """
    API xuất dữ liệu chấm công theo tháng
    
    Body: JSON
    {
        "month": "YYYY-MM"  # Tháng cần xuất
    }
    
    Returns:
        JSON response với dữ liệu chấm công đã được tích hợp thông tin từ reports
    """
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        if not data:
            return jsonify({
                "status": "error",
                "message": "Invalid JSON data"
            }), 400
            
        month = data.get("month")  # Định dạng YYYY-MM

        if not month:
            return jsonify({
                "status": "error", 
                "message": "Thiếu thông tin month"
            }), 400
        
        # Validate month format
        try:
            datetime.strptime(f"{month}-01", "%Y-%m-%d")
        except ValueError:
            return jsonify({
                "status": "error",
                "message": "Invalid month format. Use YYYY-MM"
            }), 400
        
        # Lấy danh sách nhân viên từ collection users
        users = list(user_collection.find({"active": True}))  # Chỉ lấy user active
        
        # Tạo mapping room_id → room_name
        room_mapping = {}
        unique_room_ids = list(set([user.get("room_id") for user in users if user.get("room_id")]))
        
        # Lấy thông tin phòng ban từ rooms collection (nếu có) hoặc tạo mapping mặc định
        try:
            # Nếu có rooms collection riêng
            for room_id in unique_room_ids:
                if room_id:
                    # Thử tìm trong rooms collection trước
                    room_doc = room_collection.find_one({"room_id": room_id})
                    if room_doc:
                        room_mapping[room_id] = room_doc.get("room_name", f"Phòng {room_id}")
                    else:
                        # Fallback: tạo tên phòng mặc định
                        room_mapping[room_id] = f"Phòng {room_id}"
        except:
            # Nếu không có rooms collection, tạo mapping mặc định
            for room_id in unique_room_ids:
                if room_id:
                    room_mapping[room_id] = f"Phòng {room_id}"
        
        LOGGER.info(f"Room mapping created: {room_mapping}")
        
        results = []

        # Map ngày trong tuần sang tiếng Việt
        weekday_map = {
            0: "Thứ 2",
            1: "Thứ 3", 
            2: "Thứ 4",
            3: "Thứ 5",
            4: "Thứ 6",
            5: "Thứ 7",
            6: "Chủ nhật"
        }

        # Tính toán khoảng thời gian của tháng
        start_date = datetime.strptime(f"{month}-01", "%Y-%m-%d")
        # Tính ngày cuối tháng
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1, day=1) - timedelta(days=1)

        # Lấy tất cả approved reports trong tháng này để tối ưu query
        month_start_str = start_date.strftime("%Y-%m-%d")
        month_end_str = end_date.strftime("%Y-%m-%d")
        
        approved_reports = list(reports_collection.find({
            "status": "approved",
            "date": {
                "$gte": month_start_str,
                "$lte": month_end_str
            },
            "report_type": {"$in": ["machine_error", "leave_request"]}
        }))
        
        # Tạo mapping report theo user_id và date để tối ưu lookup
        reports_map = {}
        for report in approved_reports:
            key = f"{report['user_id']}_{report['date']}"
            if key not in reports_map:
                reports_map[key] = []
            reports_map[key].append(report)

        LOGGER.info(f"Found {len(approved_reports)} approved reports for month {month}")

        # Duyệt qua từng nhân viên
        for user in users:
            user_id = user["user_id"]
            full_name = user["name"]
            room_id = user.get("room_id")
            room_name = room_mapping.get(room_id, "Chưa xác định") if room_id else "Chưa xác định"

            # Lặp qua từng ngày trong tháng
            current_date = start_date
            while current_date <= end_date:
                # Chỉ xét từ thứ 2 đến thứ 6
                if current_date.weekday() < 5:
                    date_str = current_date.strftime("%Y-%m-%d")
                    weekday_str = weekday_map[current_date.weekday()]

                    # Lấy dữ liệu attendance
                    attendance_data = all_room_logs.find_one({
                        "date": date_str,
                        "user_id": user_id
                    })
                    
                    # Tính giờ công và ghi chú với error handling
                    try:
                        check_in, check_out, total_hours, note, kpi_deduction = calculate_work_hours_new(attendance_data)
                    except Exception as calc_error:
                        LOGGER.warning(f"Error calculating work hours for user {user_id} on {date_str}: {str(calc_error)}")
                        # Fallback values
                        check_in = attendance_data.get("check_in_time", "") if attendance_data else ""
                        check_out = attendance_data.get("check_out_time", "") if attendance_data else ""
                        total_hours = "0:00"
                        note = f"Lỗi tính toán: {str(calc_error)}"
                        kpi_deduction = 0

                    # Kiểm tra và thêm ghi chú từ approved reports
                    report_key = f"{user_id}_{date_str}"
                    if report_key in reports_map:
                        report_notes = []
                        for report in reports_map[report_key]:
                            if report["report_type"] == "machine_error":
                                error_type = report.get("error_type", "other")
                                error_desc = machine_error_report_type.get(error_type, "Lỗi không xác định")
                                report_note = f"Lỗi máy: {error_desc}"
                            elif report["report_type"] == "leave_request":
                                request_type = report.get("request_type", "other")
                                request_desc = leave_request_report_type.get(request_type, "Khác")
                                report_note = f"Xin nghỉ: {request_desc}"
                            else:
                                continue
                            
                            # Thêm description từ report nếu có
                            if report.get("description"):
                                report_note += f" - {report['description']}"
                            
                            report_notes.append(report_note)
                        
                        # Thêm report notes vào ghi chú hiện có
                        if report_notes:
                            if note and note.strip():
                                note += " | " + " | ".join(report_notes)
                            else:
                                note = " | ".join(report_notes)

                    # Thêm vào kết quả
                    results.append({
                        "ID": user_id,
                        "Tên nhân viên": full_name,
                        "Phòng ban": room_name,  # ← Field mới
                        "Ngày": date_str,
                        "Thứ": weekday_str,
                        "Thời gian vào": check_in,
                        "Thời gian ra": check_out,
                        "Tổng giờ công": total_hours,
                        "Phạt tiền": kpi_deduction,
                        "Ghi chú": note
                    })

                # Sang ngày tiếp theo
                current_date += timedelta(days=1)

        # Tính toán thống kê phòng ban
        room_stats = {}
        for result in results:
            room = result["Phòng ban"]
            if room not in room_stats:
                room_stats[room] = 0
            room_stats[room] += 1

        # Trả về theo format REST API chuẩn
        return jsonify({
            "status": "success",
            "message": "Attendance data retrieved successfully",
            "data": {
                "month": month,
                "total_records": len(results),
                "total_users": len(users),
                "working_days": len([d for d in results if d["ID"] == users[0]["user_id"]]) if users else 0,
                "approved_reports_integrated": len(approved_reports),
                "rooms": list(room_mapping.values()),  # ← Danh sách phòng ban
                "room_stats": room_stats,  # ← Thống kê theo phòng
                "attendance_records": results
            }
        }), 200

    except Exception as e:
        LOGGER.error(f"Error in export_attendance: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Failed to export attendance data: {str(e)}"
        }), 500
    

# ------------------------------------------------------------
@attendance_bp.route('/generate_excel', methods=['POST'])
@login_required
@role_required('generate_excel')
def generate_excel():
    try:
        # Lấy thông tin từ yêu cầu API
        data = request.json
        edited_data = data.get("data", [])
        month_str = data.get("month", "")  # YYYY-MM
        
        if not edited_data:
            return jsonify({"error": "Không có dữ liệu để xuất"}), 400
        
        # Tạo đường dẫn tệp Excel động
        month_number = month_str.split("-")[1] if "-" in month_str else ""
        file_suffix = f"t{int(month_number)}" if month_number.isdigit() else "export"
        
        # Tạo tệp tạm thời
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, prefix=f'attendance_{file_suffix}_') as temp:
            temp_path = temp.name
        
        # Đảm bảo thứ tự cột đúng
        column_order = [
            "ID", 
            "Tên nhân viên", 
            "Ngày", 
            "Thứ", 
            "Thời gian vào", 
            "Thời gian ra", 
            "Tổng giờ công", 
            "Phạt tiền", 
            "Ghi chú"
        ]
        
        # Tạo DataFrame và sắp xếp cột theo thứ tự đã định
        df = pd.DataFrame(edited_data)
        
        # Đảm bảo tất cả các cột tồn tại (thêm cột thiếu nếu cần)
        for col in column_order:
            if col not in df.columns:
                df[col] = ""
        
        # Sắp xếp lại các cột theo thứ tự chỉ định
        df = df[column_order]
        
        # Xuất ra Excel với định dạng đẹp
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Chấm Công')
            
            # Lấy workbook và worksheet để định dạng
            workbook = writer.book
            worksheet = writer.sheets['Chấm Công']
            
            # Tạo định dạng cho tiêu đề
            header_font = Font(name='Calibri Light', bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Tạo định dạng cho nội dung
            content_font = Font(name='Calibri Light')
            content_alignment = Alignment(horizontal='center', vertical='center')
            
            # Áp dụng định dạng cho tiêu đề
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Áp dụng định dạng cho nội dung
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = content_font
                    cell.alignment = content_alignment
            
            # Tự động điều chỉnh chiều rộng cột
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Tạo đường viền cho bảng
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            # Áp dụng đường viền cho tất cả các ô có dữ liệu
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                        min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
        
        # Trả về đường dẫn đến tệp tạm thời để tải xuống
        return jsonify({
            "message": "Excel generated successfully", 
            "file": temp_path
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ------------------------------------------------------------
@attendance_bp.route('/download', methods=['GET'])
@login_required
@role_required('download')
def download_file():
    file_path = request.args.get('file')
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({"error": "File không tồn tại"}), 404
    
    # Lấy tên tệp từ đường dẫn
    filename = os.path.basename(file_path)
    
    # Gửi tệp về client
    response = send_file(file_path, as_attachment=True, download_name=filename)
    
    # Xóa tệp tạm thời sau khi gửi
    @response.call_on_close
    def cleanup():
        try:
            os.remove(file_path)
        except:
            pass
    
    return response


# Cấu hình thời gian mới
TIME_CONFIG = {
    'WORK_START_TIME': '08:00:00',      # Giờ bắt đầu làm việc chuẩn
    'LATE_THRESHOLD': '08:10:00',       # Giờ bắt đầu tính muộn nhẹ
    'VERY_LATE_THRESHOLD': '08:30:00',  # Giờ bắt đầu tính muộn nặng
    'AFTERNOON_START': '14:00:00',      # Giờ bắt đầu làm việc buổi chiều
    'WORK_END_TIME': '17:30:00',        # Giờ kết thúc làm việc
    'WORK_HOURS_REQUIRED': 8,           # Số giờ làm việc yêu cầu
    'LUNCH_BREAK_HOURS': 1.5,           # Thời gian nghỉ trưa
    'LATE_FINE': 30,                    # Tiền phạt đi muộn nhẹ
    'VERY_LATE_FINE': 50                # Tiền phạt đi muộn nặng
}

WEEKLY_LATE_LIMIT = 3
WEEKLY_LATE_FINE = 100


def normalize_time_format(time_str):
    """
    Chuẩn hóa format thời gian về HH:MM:SS
    Input: "08:00" hoặc "08:00:00"
    Output: "08:00:00"
    """
    if not time_str:
        return time_str
    
    # Nếu đã có seconds thì giữ nguyên
    if len(time_str.split(':')) == 3:
        return time_str
    
    # Nếu chỉ có HH:MM thì thêm :00
    if len(time_str.split(':')) == 2:
        return f"{time_str}:00"
    
    return time_str


def calculate_work_hours_new(attendance_data):
    """
    Tính tổng giờ công trong ngày, ghi chú và phạt tiền theo logic mới
    """
    if not attendance_data:  # Không có dữ liệu
        return None, None, None, "Nghỉ", 0
    
    check_in_time = attendance_data.get("check_in_time")
    check_out_time = attendance_data.get("check_out_time")
    
    # Chuẩn hóa format thời gian
    check_in_time = normalize_time_format(check_in_time)
    check_out_time = normalize_time_format(check_out_time)
    
    if not check_in_time or not check_out_time:
        return check_in_time, check_out_time, None, "Nghỉ", 0
    
    try:
        # Chuyển đổi định dạng thời gian
        from datetime import datetime
        check_in_dt = datetime.strptime(check_in_time, "%H:%M:%S")
        check_out_dt = datetime.strptime(check_out_time, "%H:%M:%S")
        
        # Tính tổng thời gian từ check-in đến check-out (bao gồm cả nghỉ trưa)
        total_seconds = (check_out_dt - check_in_dt).total_seconds()
        total_hours_with_lunch = total_seconds / 3600
        
        # Tính giờ làm việc thực tế (trừ giờ nghỉ trưa)
        if total_hours_with_lunch > TIME_CONFIG['LUNCH_BREAK_HOURS']:
            actual_work_hours = total_hours_with_lunch - TIME_CONFIG['LUNCH_BREAK_HOURS']
        else:
            actual_work_hours = total_hours_with_lunch
        
        actual_work_hours = round(actual_work_hours, 2)
        
        # Thời gian chuẩn để so sánh
        late_threshold = datetime.strptime(TIME_CONFIG['LATE_THRESHOLD'], "%H:%M:%S")
        very_late_threshold = datetime.strptime(TIME_CONFIG['VERY_LATE_THRESHOLD'], "%H:%M:%S")
        afternoon_start = datetime.strptime(TIME_CONFIG['AFTERNOON_START'], "%H:%M:%S")
        
        # Kiểm tra các điều kiện
        is_late = check_in_dt > late_threshold
        is_very_late = check_in_dt > very_late_threshold
        missing_morning = check_in_dt > afternoon_start  # Thiếu công sáng
        missing_afternoon = check_out_dt < afternoon_start  # Thiếu công chiều
        
        # THAY ĐỔI: Kiểm tra về sớm dựa trên tổng thời gian làm việc
        # Nếu tổng thời gian check-in -> check-out < 9.5 tiếng (8h làm việc + 1.5h nghỉ trưa)
        min_total_hours_required = TIME_CONFIG['WORK_HOURS_REQUIRED'] + TIME_CONFIG['LUNCH_BREAK_HOURS']  # 9.5 tiếng
        is_early_leave = total_hours_with_lunch < min_total_hours_required
        
        # Xác định ghi chú
        notes = []
        fine_amount = 0
        
        if missing_morning:
            notes.append("Thiếu công sáng")
        elif missing_afternoon:
            notes.append("Thiếu công chiều")
        else:
            # Chỉ tính phạt khi không thiếu công
            if is_very_late:
                late_minutes = int((check_in_dt - very_late_threshold).total_seconds() / 60)
                notes.append(f"Đi muộn nặng ({late_minutes + 20} phút)")
                fine_amount = TIME_CONFIG['VERY_LATE_FINE']
            elif is_late:
                late_minutes = int((check_in_dt - late_threshold).total_seconds() / 60)
                notes.append(f"Đi muộn ({late_minutes + 10} phút)")
                fine_amount = TIME_CONFIG['LATE_FINE']
            
            # THAY ĐỔI: Kiểm tra về sớm dựa trên tổng thời gian
            if is_early_leave:
                missing_hours = min_total_hours_required - total_hours_with_lunch
                missing_minutes = int(missing_hours * 60)
                notes.append(f"Về sớm (thiếu {missing_minutes} phút)")
        
        note = " & ".join(notes) if notes else ""
        
        # Trả về giờ làm việc thực tế (đã trừ nghỉ trưa)
        return check_in_time, check_out_time, actual_work_hours, note, fine_amount
        
    except ValueError as e:
        LOGGER.error(f"Error parsing time format: {e}")
        return check_in_time, check_out_time, None, f"Lỗi format thời gian: {str(e)}", 0


def update_attendance_for_machine_error(report):
    """
    Cập nhật attendance record dựa trên thông tin từ machine_error report
    - Chuyển đổi error_time thành thời gian cụ thể và cập nhật attendance
    """
    try:
        user_id = report["user_id"]
        date = report["date"]
        error_time = report["error_time"]
        
        # Mapping error_time thành thời gian cụ thể
        time_mapping = {
            "morning": "08:00:00",      # 8h00
            "afternoon": "17:30:00",    # 17h30
            "full_day": ["08:00:00", "17:30:00"]  # Cả 2 thời điểm
        }
        
        if error_time not in time_mapping:
            raise Exception(f"Invalid error_time: {error_time}")
        
        # Tìm attendance record hiện có
        attendance_filter = {
            "user_id": user_id,
            "date": date
        }
        
        existing_attendance = all_room_logs.find_one(attendance_filter)
        
        # Lấy thông tin user
        user_info = user_collection.find_one({"user_id": user_id})
        user_name = user_info.get("name", "Unknown User") if user_info else "Unknown User"
        
        results = []
        
        if error_time == "full_day":
            # Xử lý cả check-in và check-out
            times_to_process = [
                ("08:00:00", "check_in"),
                ("17:30:00", "check_out")
            ]
        else:
            # Xử lý một thời điểm
            time_value = time_mapping[error_time]
            time_type = "check_in" if error_time == "morning" else "check_out"
            times_to_process = [(time_value, time_type)]
        
        if existing_attendance:
            # Update attendance record hiện có
            update_data = {}
            
            for time_value, time_type in times_to_process:
                if time_type == "check_in":
                    update_data.update({
                        "check_in_time": time_value,
                        "has_been_welcome": True
                    })
                else:
                    update_data.update({
                        "check_out_time": time_value,
                        "has_been_goodbye": True
                    })
                
                results.append({
                    "action": f"updated_{time_type}",
                    "time": time_value
                })
            
            result = all_room_logs.update_one(
                attendance_filter,
                {"$set": update_data}
            )
            
            if result.modified_count > 0:
                return {
                    "user_id": user_id,
                    "date": date,
                    "error_time": error_time,
                    "updates": results,
                    "record_existed": True
                }
            else:
                raise Exception("Failed to update existing attendance record")
                
        else:
            # Tạo attendance record mới
            new_attendance = {
                "date": date,
                "user_id": user_id,
                "name": user_name,
                "timestamps": [],
                "has_been_welcome": False,
                "has_been_goodbye": False
            }
            
            for time_value, time_type in times_to_process:
                if time_type == "check_in":
                    new_attendance.update({
                        "check_in_time": time_value,
                        "check_in_image": "",
                        "has_been_welcome": True
                    })
                else:
                    new_attendance.update({
                        "check_out_time": time_value,
                        "has_been_goodbye": True
                    })
                
                results.append({
                    "action": f"created_with_{time_type}",
                    "time": time_value
                })
            
            result = all_room_logs.insert_one(new_attendance)
            
            if result.inserted_id:
                return {
                    "user_id": user_id,
                    "date": date,
                    "error_time": error_time,
                    "updates": results,
                    "record_existed": False,
                    "new_record_id": str(result.inserted_id)
                }
            else:
                raise Exception("Failed to create new attendance record")
        
    except Exception as e:
        LOGGER.error(f"Error in update_attendance_for_machine_error: {str(e)}")
        raise e